import io
import json
import openpyxl
import pytest
from datetime import datetime
from fastapi.testclient import TestClient


@pytest.fixture()
def client(monkeypatch):
    monkeypatch.setenv("REPORTING_DB_URL", "postgresql://unused")
    monkeypatch.delenv("IMPORT_TOKEN", raising=False)
    import exporter.app as app_module
    return TestClient(app_module.app), app_module


def test_import_form_renders(client):
    c, _ = client
    r = c.get("/import")
    assert r.status_code == 200
    assert "text/html" in r.headers["content-type"]
    assert "<form" in r.text and 'type="file"' in r.text


def _filled_workbook():
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    p = wb.create_sheet("2. Projects"); p["A3"], p["B3"] = "P01", "Future"
    m = wb.create_sheet("3. Monthly")
    m["A4"], m["B4"], m["E4"], m["P4"] = "P01", datetime(2026, 6, 1), 19, 30
    wb.create_sheet("4. Quarterly")
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_preview_shows_diff_and_embeds_changes(client, monkeypatch):
    c, app_module = client
    monkeypatch.setattr(app_module, "fetch_manual",
                        lambda db, ps: {("Future", "2026-06"): {"total_engineers": "18"}})
    monkeypatch.setattr(app_module, "fetch_auto_ai_users", lambda db, ps: {})
    r = c.post("/import/preview", files={"file": ("wb.xlsx", _filled_workbook())})
    assert r.status_code == 200
    assert "changed" in r.text          # total_engineers 18 -> 19
    assert "total_engineers" in r.text
    # the confirm form embeds the parsed changes as JSON
    assert "/import/commit" in r.text and "total_engineers" in r.text


def test_preview_rejects_non_workbook(client):
    c, _ = client
    r = c.post("/import/preview", files={"file": ("x.xlsx", b"not a zip")})
    assert r.status_code == 400


def test_preview_rejects_wrong_shape_workbook(client):
    c, _ = client
    wb = openpyxl.Workbook()  # a valid xlsx, but NOT the maturity template
    buf = io.BytesIO(); wb.save(buf)
    r = c.post("/import/preview", files={"file": ("wrong.xlsx", buf.getvalue())})
    assert r.status_code == 400


def test_token_guard_rejects_when_set(client, monkeypatch):
    c, _ = client
    monkeypatch.setenv("IMPORT_TOKEN", "secret")
    # preview with no token -> 401 (guard runs before the file is read)
    r = c.post("/import/preview", files={"file": ("wb.xlsx", _filled_workbook())})
    assert r.status_code == 401


def test_commit_writes_each_change(client, monkeypatch):
    c, app_module = client
    written = []
    monkeypatch.setattr(app_module, "upsert_manual_input",
                        lambda db, project, period, field, value, by: written.append(
                            (project, period, field, value, by)))
    changes = json.dumps([
        {"project": "Future", "period_key": "2026-06", "field": "total_engineers", "value": "19"},
        {"project": "Future", "period_key": "2026-Q3", "field": "g1_agents_md", "value": "Yes"},
    ])
    r = c.post("/import/commit", data={"changes": changes})
    assert r.status_code == 200
    assert len(written) == 2
    assert ("Future", "2026-06", "total_engineers", "19", "excel-import") in written
    assert "2" in r.text  # summary count


def test_commit_rejects_bad_json(client):
    c, _ = client
    r = c.post("/import/commit", data={"changes": "not json"})
    assert r.status_code == 400
