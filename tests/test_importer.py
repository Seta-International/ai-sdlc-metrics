import openpyxl
from exporter.importer import parse_manual_inputs, format_value, diff_changes


def _wb_with(projects_rows, monthly_rows=(), quarterly_rows=()):
    """Build a minimal workbook shaped like the exporter output."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    p = wb.create_sheet("2. Projects")
    for i, (pid, name) in enumerate(projects_rows):
        p[f"A{3+i}"], p[f"B{3+i}"] = pid, name
    m = wb.create_sheet("3. Monthly")
    for i, cells in enumerate(monthly_rows):
        r = 4 + i
        for col, val in cells.items():
            m[f"{col}{r}"] = val
    q = wb.create_sheet("4. Quarterly")
    for i, cells in enumerate(quarterly_rows):
        r = 4 + i
        for col, val in cells.items():
            q[f"{col}{r}"] = val
    return wb


def test_format_value_integral_float_has_no_decimal():
    assert format_value(18.0) == "18"
    assert format_value(0.55) == "0.55"
    assert format_value("Yes") == "Yes"
    assert format_value(3) == "3"


def test_parse_reads_only_manual_monthly_cells():
    from datetime import datetime
    wb = _wb_with(
        projects_rows=[("P01", "Future")],
        monthly_rows=[{"A": "P01", "B": datetime(2026, 6, 1),
                       "E": 18, "O": 45, "P": 30, "S": 0.55,
                       "F": 999, "G": 999}],  # F/G are AUTO cols — must be ignored
    )
    got = parse_manual_inputs(wb)
    assert {"project": "Future", "period_key": "2026-06",
            "field": "total_engineers", "value": "18"} in got
    assert {"project": "Future", "period_key": "2026-06",
            "field": "coverage_ai", "value": "0.55"} in got
    # auto cols never appear as fields
    assert not any(c["field"] in ("ai_prs", "total_prs") for c in got)
    assert len(got) == 4  # E,O,P,S only


def test_parse_reads_quarterly_flags_and_skips_blanks():
    wb = _wb_with(
        projects_rows=[("P01", "Future")],
        quarterly_rows=[{"A": "P01", "B": "2026-Q3", "C": "Yes", "D": "No",
                         # column AF (index for evidence) left blank -> skipped
                         }],
    )
    got = parse_manual_inputs(wb)
    assert {"project": "Future", "period_key": "2026-Q3",
            "field": "g1_agents_md", "value": "Yes"} in got
    assert {"project": "Future", "period_key": "2026-Q3",
            "field": "g2_ai_policy", "value": "No"} in got
    # only the two filled flags, nothing for blank columns
    assert len(got) == 2


def test_diff_classifies_new_changed_unchanged():
    parsed = [
        {"project": "Future", "period_key": "2026-06", "field": "total_engineers", "value": "19"},
        {"project": "Future", "period_key": "2026-06", "field": "cost_actual", "value": "30"},
        {"project": "Future", "period_key": "2026-Q3", "field": "g1_agents_md", "value": "Yes"},
    ]
    current = {("Future", "2026-06"): {"total_engineers": "18", "cost_actual": "30"}}
    got = diff_changes(parsed, current)
    by_field = {(c["period_key"], c["field"]): c for c in got}
    assert by_field[("2026-06", "total_engineers")]["status"] == "changed"
    assert by_field[("2026-06", "total_engineers")]["old"] == "18"
    assert by_field[("2026-06", "cost_actual")]["status"] == "unchanged"
    assert by_field[("2026-Q3", "g1_agents_md")]["status"] == "new"
    assert by_field[("2026-Q3", "g1_agents_md")]["old"] is None
