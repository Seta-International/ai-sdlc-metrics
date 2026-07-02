"""The committed EN template must match what the specs generate."""
import openpyxl
import pytest

from exporter.template import DST, build_workbook


@pytest.fixture(scope="module")
def generated():
    return build_workbook()


@pytest.fixture(scope="module")
def committed():
    return openpyxl.load_workbook(DST)


def test_sheet_names_and_order(generated, committed):
    assert generated.sheetnames == committed.sheetnames


def test_every_cell_matches(generated, committed):
    mismatches = []
    for name in committed.sheetnames:
        ws_c, ws_g = committed[name], generated[name]
        rows = max(ws_c.max_row, ws_g.max_row)
        cols = max(ws_c.max_column, ws_g.max_column)
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                vc = ws_c.cell(row=row, column=col).value
                vg = ws_g.cell(row=row, column=col).value
                if vc != vg:
                    mismatches.append(f"{name}!{ws_c.cell(row=row, column=col).coordinate}:"
                                      f" committed={vc!r} generated={vg!r}")
    assert not mismatches, "\n".join(mismatches[:40]) + f"\n({len(mismatches)} total)"


def test_layout_extras_match(generated, committed):
    for name in committed.sheetnames:
        ws_c, ws_g = committed[name], generated[name]
        assert ws_g.freeze_panes == ws_c.freeze_panes, name
        assert {str(r) for r in ws_g.merged_cells.ranges} == \
               {str(r) for r in ws_c.merged_cells.ranges}, name
        dv = lambda ws: {(d.type, d.formula1, str(d.sqref))
                         for d in ws.data_validations.dataValidation}
        assert dv(ws_g) == dv(ws_c), name


def test_charts_present(generated):
    assert len(generated["8. Dashboard-Project"]._charts) == 2
    assert len(generated["9. Dashboard-Portfolio"]._charts) == 1


def test_fill_columns_have_headers(generated):
    from exporter.workbook import SHEET3_MANUAL_COLS, SHEET3_METRIC_COLS
    ws = generated["3. Monthly"]
    for col in {**SHEET3_METRIC_COLS, **SHEET3_MANUAL_COLS}:
        assert isinstance(ws[f"{col}3"].value, str) and ws[f"{col}3"].value, col


def test_fill_workbook_runs_on_generated_template(generated, tmp_path):
    from datetime import date
    from exporter.workbook import fill_workbook
    path = tmp_path / "template.xlsx"
    generated.save(path)
    wb = fill_workbook(path, ["Future"], [], [
        {"project": "Future", "period_key": "2026-06",
         "period_start": date(2026, 6, 1), "ai_prs": 20, "total_prs": 50},
    ], {})
    assert wb["2. Projects"]["B3"].value == "Future"
    assert float(wb["3. Monthly"]["F4"].value) == 20.0
