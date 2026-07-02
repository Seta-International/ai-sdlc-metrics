from pathlib import Path
import openpyxl
import pytest
from exporter.build_template import DST, _LITERAL, _VN

pytestmark = pytest.mark.skipif(not DST.exists(), reason="EN template not built yet")


def test_template_fully_english():
    wb = openpyxl.load_workbook(DST)
    offenders = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str):
                    continue
                if v.startswith("="):  # also scan formula string literals
                    if any(_VN.search(s) for s in _LITERAL.findall(v)):
                        offenders.append(f"{ws.title}!{c.coordinate}")
                elif _VN.search(v):
                    offenders.append(f"{ws.title}!{c.coordinate}")
    assert offenders == []


def test_template_structure():
    wb = openpyxl.load_workbook(DST)
    assert "1. Guide" in wb.sheetnames and "3. Monthly" in wb.sheetnames
    m = wb["3. Monthly"]
    assert m["A4"].value is None            # sample data cleared
    assert str(m["C4"].value).startswith("=")  # quarter formula kept
    q = wb["4. Quarterly"]
    assert q["A4"].value is None
    assert str(q["AH4"].value).startswith("=")


def test_template_has_charts():
    wb = openpyxl.load_workbook(DST)
    assert len(wb["8. Dashboard-Project"]._charts) == 2
    assert len(wb["9. Dashboard-Portfolio"]._charts) == 1
