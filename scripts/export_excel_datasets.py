#!/usr/bin/env python3
"""
Export DB data to per-dataset Excel workbooks — one xlsx per dataset folder.
Each file starts with the LEGEND & SUMMARY sheet copied verbatim from the
original mock_data.xlsx, followed by data sheets queried from the DB.

Produces: datasets/output/{folder}.xlsx  (7 files)

Dependencies:
    pip3 install psycopg2-binary openpyxl pandas
"""
import copy
import psycopg2
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

DB = dict(host="localhost", port=55432, dbname="hackathon", user="postgres", password="postgres")
SCRIPT_DIR = Path(__file__).parent
DATASETS_DIR = SCRIPT_DIR / "../datasets"


def fetch(cur, sql):
    cur.execute(sql)
    cols = [d[0] for d in cur.description]
    df = pd.DataFrame(cur.fetchall(), columns=cols)
    for col in df.select_dtypes(include=["datetimetz"]).columns:
        df[col] = df[col].dt.tz_localize(None)
    return df


def copy_sheet(src_ws, dst_wb, dst_name):
    """Copy a worksheet (values, merges, col widths) into dst_wb."""
    dst_ws = dst_wb.create_sheet(title=dst_name)
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst_cell.font      = copy.copy(cell.font)
                dst_cell.fill      = copy.copy(cell.fill)
                dst_cell.border    = copy.copy(cell.border)
                dst_cell.alignment = copy.copy(cell.alignment)
                dst_cell.number_format = cell.number_format
    for mc in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(mc))
    for col_letter, cd in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = cd.width
    for row_dim, rd in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_dim].height = rd.height
    return dst_ws


def write_xlsx(out_dir, folder, sheets):
    """Write one xlsx: legend sheet (from original) + data sheets."""
    from datetime import date
    from openpyxl.styles import Font, PatternFill, Alignment

    path = out_dir / f"{folder}.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)

    original = DATASETS_DIR / folder / "originals" / "mock_data.xlsx"
    src_wb = openpyxl.load_workbook(str(original), data_only=True)
    legend_ws = src_wb[src_wb.sheetnames[0]]
    legend_name = legend_ws.title

    wb = Workbook()
    wb.remove(wb.active)           # remove default empty sheet
    legend_dst = copy_sheet(legend_ws, wb, legend_name)
    src_wb.close()

    # Update row counts in the legend sheet if they appear as integers near a sheet name
    sheet_name_to_count = {name: len(df) for name, df in sheets}
    for row in legend_dst.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() in sheet_name_to_count:
                # Look right for an integer cell to update
                for offset in range(1, 5):
                    neighbor = legend_dst.cell(row=cell.row, column=cell.column + offset)
                    if isinstance(neighbor.value, (int, float)) and neighbor.value > 0:
                        neighbor.value = sheet_name_to_count[cell.value.strip()]
                        break

    # Append DB schema summary block at bottom of legend sheet
    last_row = legend_dst.max_row + 2
    header_fill = PatternFill("solid", fgColor="2F4F4F")
    header_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="F0F4F4")

    hdr = legend_dst.cell(row=last_row, column=1,
                           value=f"DB Schema Summary — exported {date.today()} from PostgreSQL 16")
    hdr.font = Font(bold=True, size=12)
    last_row += 1

    labels = ["Sheet Name", "Rows Exported", "Columns"]
    for col_idx, label in enumerate(labels, 1):
        c = legend_dst.cell(row=last_row, column=col_idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    last_row += 1

    for i, (sheet_name, df) in enumerate(sheets):
        fill = alt_fill if i % 2 == 0 else None
        vals = [sheet_name, len(df), ", ".join(df.columns.tolist()[:8])]
        for col_idx, val in enumerate(vals, 1):
            c = legend_dst.cell(row=last_row, column=col_idx, value=val)
            if fill:
                c.fill = fill
        last_row += 1

    # Set column widths for the summary block
    legend_dst.column_dimensions["A"].width = max(
        legend_dst.column_dimensions["A"].width or 20,
        max((len(n) for n, _ in sheets), default=20) + 2
    )
    legend_dst.column_dimensions["B"].width = 15
    legend_dst.column_dimensions["C"].width = 60

    print(f"    ✓ {legend_name} (legend — updated row counts + DB schema summary)")

    # append data sheets
    for sheet_name, df in sheets:
        ws = wb.create_sheet(title=sheet_name[:31])
        # header row
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # data rows
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        # auto-fit column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(
                len(str(col_name)),
                df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
        print(f"    ✓ {sheet_name} ({len(df)} rows)")

    wb.save(str(path))
    print(f"  → {path}")


# ---------------------------------------------------------------------------
# DS01 — PMO Project Plan Review
# ---------------------------------------------------------------------------
def export_ds01(cur, out_dir, name):
    sheets = [
        ("DS01_Project_Plan", fetch(cur, """
            SELECT p.project_code AS "Project_ID",
                   p.name AS "Project_name",
                   pt.task_code AS "Task_ID",
                   pt.task_name AS "Task_name",
                   regexp_replace(e.emp_code, '^EMP-0(\\d{3})$', 'EMP-\\1') AS "Assignee_id",
                   pt.start_date AS "Start_date",
                   pt.end_date AS "End_date",
                   pt.effort_days AS "Effort_days",
                   pt.percent_complete AS "Percent_complete",
                   pt.status AS "Status",
                   pt.is_milestone AS "Milestone_flag",
                   NULL::text AS "Dependencies",
                   pt.phase AS "Phase",
                   pt.risk_note AS "Risk_note"
            FROM pmo.plan_task pt
            JOIN pmo.plan pl ON pl.plan_id = pt.plan_id
            JOIN core.project p ON p.project_id = pl.project_id
            LEFT JOIN core.employee e ON e.employee_id = pt.assignee_id
            ORDER BY p.project_code, pt.start_date
        """)),
        ("DS02_PMO_Standard_Template", fetch(cur, """
            SELECT pt.template_code AS "Template_ID",
                   pt.name AS "Template_name",
                   pt.version AS "Version",
                   pt.effective_date AS "Effective_date",
                   tc.component_code AS "Component_ID",
                   tc.section_code AS "Section_code",
                   tc.component_name AS "Component_name",
                   tc.is_required AS "Required",
                   tc.validation_rule AS "Validation_rule",
                   tc.weight AS "Weight"
            FROM pmo.plan_template pt
            JOIN pmo.template_component tc ON tc.plan_template_id = pt.plan_template_id
            ORDER BY pt.template_code, tc.component_code
        """)),
        ("DS03_Resource_Allocation", fetch(cur, """
            SELECT regexp_replace(e.emp_code, '^EMP-0(\\d{3})$', 'EMP-\\1') AS "Member_ID",
                   p.project_code AS "Project_ID",
                   r.role_code AS "Role",
                   ra.allocation_pct AS "Allocation_pct",
                   ra.start_date AS "Start_date",
                   ra.end_date AS "End_date",
                   vbr.busy_rate AS "Busy_rate"
            FROM pmo.resource_allocation ra
            JOIN core.employee e ON e.employee_id = ra.employee_id
            JOIN core.project p ON p.project_id = ra.project_id
            JOIN core.role r ON r.role_id = ra.role_id
            LEFT JOIN pmo.v_member_busy_rate vbr ON vbr.employee_id = ra.employee_id
            WHERE p.project_code ~ '^PRJ-'
            ORDER BY ra.start_date, e.emp_code
        """)),
        ("DS04_Velocity_History", fetch(cur, """
            SELECT p.project_code AS "Project_ID",
                   ptype.name AS "Project_type",
                   vh.sprint_no AS "Sprint_no",
                   vh.sprint_duration_days AS "Sprint_duration_days",
                   vh.planned_points AS "Planned_points",
                   vh.completed_points AS "Completed_points",
                   vh.velocity_ratio AS "Velocity_ratio",
                   vh.team_size AS "Team_size",
                   vh.outcome AS "Outcome"
            FROM pmo.velocity_history vh
            JOIN core.project p ON p.project_id = vh.project_id
            JOIN core.project_type ptype ON ptype.project_type_id = p.project_type_id
            ORDER BY p.project_code, vh.sprint_no
        """)),
        ("DS05_Historical_Projects", fetch(cur, """
            SELECT p.project_code AS "Historical_project_id",
                   ptype.name AS "Project_type",
                   hb.team_size AS "Team_size",
                   hb.duration_days AS "Duration_days",
                   hb.planned_duration_days AS "Planned_duration_days",
                   hb.total_effort_days AS "Total_effort_days",
                   hb.total_budget_scaled AS "Total_budget_scaled",
                   hb.avg_velocity_ratio AS "Avg_velocity_ratio",
                   hb.risk_count AS "Risk_count",
                   hb.key_risks AS "Key_risks",
                   hb.pmo_standard_ver AS "PMO_standard_ver",
                   hb.final_outcome AS "Final_outcome",
                   hb.is_outlier AS "Is_outlier"
            FROM pmo.historical_benchmark hb
            JOIN core.project p ON p.project_id = hb.project_id
            JOIN core.project_type ptype ON ptype.project_type_id = p.project_type_id
            ORDER BY p.project_code
        """)),
        ("DS06_Plan_Section_Check", fetch(cur, """
            SELECT 'CHK-' || LPAD(psc.plan_section_check_id::text, 3, '0') AS "Check_ID",
                   pl.plan_code AS "Plan_ID",
                   tc.component_code AS "Component_ID",
                   NULLIF(psc.custom_name, '') AS "Custom_name",
                   psc.status AS "Status",
                   NULLIF(psc.note, '') AS "Note"
            FROM pmo.plan_section_check psc
            JOIN pmo.plan pl ON pl.plan_id = psc.plan_id
            LEFT JOIN pmo.template_component tc
                ON tc.template_component_id = psc.template_component_id
            ORDER BY pl.plan_code, tc.component_code NULLS LAST
        """)),
        ("DS07_Project_Plan_Summary", fetch(cur, """
            SELECT pl.plan_code AS "Plan_ID",
                   p.project_code AS "Project_ID",
                   p.name AS "Project_name",
                   pl.plan_set AS "Plan_set",
                   ps.effort_md AS "Effort_MD",
                   pl.planned_duration_months AS "Duration_months",
                   ps.velocity_md_month AS "Velocity_MD_month",
                   pl.team_size_planned AS "Team_size",
                   pl.registered_risk_count AS "Risk_count",
                   pl.top_risk_score AS "Top_risk_score",
                   pl.thi_pct AS "THI_pct",
                   pl.peak_role_busy_rate_pct AS "Peak_role_busy_rate_pct",
                   pl.on_time_history_pct AS "On_time_history_pct",
                   pl.feasibility_status AS "Feasibility_status"
            FROM pmo.plan pl
            JOIN core.project p ON p.project_id = pl.project_id
            LEFT JOIN pmo.v_plan_summary ps ON ps.plan_id = pl.plan_id
            ORDER BY pl.plan_code
        """)),
        ("DS08_Role_Capacity", fetch(cur, """
            SELECT 'CAP-' || LPAD(rc.role_capacity_id::text, 2, '0') AS "Capacity_ID",
                   r.name AS "Role",
                   rc.headcount AS "Headcount",
                   rc.capacity_md_month AS "Capacity_MD_month",
                   rc.busy_rate_pct AS "Busy_rate_pct",
                   rc.available_md_month AS "Available_MD_month",
                   rc.note AS "Note"
            FROM pmo.role_capacity rc
            JOIN core.role r ON r.role_id = rc.role_id
            ORDER BY r.role_code
        """)),
        ("REF_Member_Master", fetch(cur, """
            SELECT regexp_replace(e.emp_code, '^EMP-0(\\d{3})$', 'EMP-\\1') AS "Member_ID",
                   e.full_name AS "Full_name",
                   e.position_title AS "Role_title",
                   d.name AS "Department",
                   e.employment_type AS "Employment",
                   e.std_hours_week AS "Std_hours_week"
            FROM core.employee e
            JOIN core.department d ON d.department_id = e.department_id
            WHERE e.employee_id IN (
              SELECT DISTINCT ra.employee_id FROM pmo.resource_allocation ra
              JOIN core.project p ON p.project_id = ra.project_id
              WHERE p.project_code ~ '^PRJ-'
              UNION
              SELECT pt.assignee_id FROM pmo.plan_task pt WHERE pt.assignee_id IS NOT NULL
              UNION
              SELECT p.pm_id FROM core.project p WHERE p.pm_id IS NOT NULL
                AND p.project_code ~ '^PRJ-'
            )
            ORDER BY e.emp_code
        """)),
        ("REF_Project_Master", fetch(cur, """
            SELECT p.project_code AS "Project_ID",
                   p.name AS "Project_name",
                   ptype.name AS "Project_type",
                   p.status AS "Status",
                   p.is_historical AS "Is_historical"
            FROM core.project p
            JOIN core.project_type ptype ON ptype.project_type_id = p.project_type_id
            WHERE p.project_id IN (
              SELECT project_id FROM pmo.plan
              UNION SELECT project_id FROM pmo.velocity_history
              UNION SELECT project_id FROM pmo.historical_benchmark
              UNION SELECT project_id FROM pmo.resource_allocation
                WHERE project_id IN (
                  SELECT project_id FROM core.project WHERE project_code ~ '^PRJ-'
                )
            )
            ORDER BY p.project_code
        """)),
        ("REF_KPI_Norms", fetch(cur, """
            SELECT mn.norm_code AS "Norm_ID",
                   mn.metric AS "Metric",
                   mn.formula AS "Formula",
                   MAX(CASE WHEN mnt.rag = 'Green' THEN mnt.rule_expr END) AS "Green",
                   MAX(CASE WHEN mnt.rag = 'Yellow' THEN mnt.rule_expr END) AS "Yellow",
                   MAX(CASE WHEN mnt.rag = 'Red' THEN mnt.rule_expr END) AS "Red",
                   mn.used_for AS "Used_for"
            FROM core.metric_norm mn
            LEFT JOIN core.metric_norm_threshold mnt
                ON mnt.metric_norm_id = mn.metric_norm_id
            GROUP BY mn.norm_code, mn.metric, mn.formula, mn.used_for
            ORDER BY mn.norm_code
        """)),
    ]
    write_xlsx(out_dir, name, sheets)


# ---------------------------------------------------------------------------
# DS02 — PMO Timesheet Monitoring
# ---------------------------------------------------------------------------
def export_ds02(cur, out_dir, name):
    sheets = [
        ("DS01_Resource_Allocation", fetch(cur, """
            SELECT regexp_replace(e.emp_code, '^EMP-0(\\d{3})$', 'EMP-\\1') AS "Member_ID",
                   p.project_code AS "Project_ID",
                   r.role_code AS "Role",
                   ra.allocation_pct AS "Allocation_pct",
                   ra.start_date AS "Start_date",
                   ra.end_date AS "End_date",
                   ROUND(ra.allocation_pct * e.std_hours_week) AS "Weekly_planned_hours"
            FROM pmo.resource_allocation ra
            JOIN core.employee e ON e.employee_id = ra.employee_id
            JOIN core.project p ON p.project_id = ra.project_id
            JOIN core.role r ON r.role_id = ra.role_id
            WHERE p.project_code ~ '^PRJ-'
            ORDER BY e.emp_code
        """)),
        ("DS02_Timesheet_Log", fetch(cur, """
            SELECT regexp_replace(e.emp_code, '^EMP-0(\\d{3})$', 'EMP-\\1') AS "Member_ID",
                   p.project_code AS "Project_ID",
                   tl.work_date AS "Work_date",
                   tl.logged_hours AS "Logged_hours",
                   tl.log_category AS "Log_category",
                   tl.task_ref AS "Task_ref"
            FROM pmo.timesheet_log tl
            JOIN core.employee e ON e.employee_id = tl.employee_id
            LEFT JOIN core.project p ON p.project_id = tl.project_id
            ORDER BY tl.work_date, e.emp_code
        """)),
        ("DS03_Overbook_Idle_Config", fetch(cur, """
            SELECT config_code AS "Config_ID",
                   rule_name AS "Rule_name",
                   overbook_threshold AS "Overbook_threshold",
                   overbook_red_threshold AS "Overbook_red_threshold",
                   idle_threshold AS "Idle_threshold",
                   mismatch_pct_threshold AS "Mismatch_pct_threshold",
                   ot_max_hours_per_week AS "OT_max_hours_per_week",
                   effective_date AS "Effective_date"
            FROM pmo.overbook_idle_config
        """)),
        ("DS04_Leave_Holiday_Records", fetch(cur, """
            SELECT lr.leave_record_code AS "Record_ID",
                   regexp_replace(e.emp_code, '^EMP-0(\\d{3})$', 'EMP-\\1') AS "Member_ID",
                   lr.leave_date AS "Leave_date",
                   lr.leave_type AS "Leave_type",
                   lr.approved AS "Approved",
                   lr.duration_days AS "Duration_days",
                   lr.note AS "Note"
            FROM pmo.leave_record lr
            LEFT JOIN core.employee e ON e.employee_id = lr.employee_id
            ORDER BY lr.leave_date, e.emp_code NULLS LAST
        """)),
        ("DS05_Project_Master", fetch(cur, """
            SELECT p.project_code AS "Project_ID",
                   p.name AS "Project_name",
                   a.account_code AS "Account_ID",
                   ptype.name AS "Project_type",
                   p.status AS "Status",
                   regexp_replace(pm.emp_code, '^EMP-0(\\d{3})$', 'EMP-\\1') AS "PM_ID",
                   p.start_date AS "Start_date",
                   p.planned_end_date AS "End_date"
            FROM core.project p
            JOIN core.account a ON a.account_id = p.account_id
            JOIN core.project_type ptype ON ptype.project_type_id = p.project_type_id
            LEFT JOIN core.employee pm ON pm.employee_id = p.pm_id
            WHERE p.project_id IN (
              SELECT project_id FROM pmo.plan
              UNION SELECT project_id FROM pmo.velocity_history
              UNION SELECT project_id FROM pmo.historical_benchmark
              UNION SELECT project_id FROM pmo.resource_allocation
                WHERE project_id IN (
                  SELECT project_id FROM core.project WHERE project_code ~ '^PRJ-'
                )
            )
            ORDER BY p.project_code
        """)),
        ("DS06_Member_Master", fetch(cur, """
            SELECT regexp_replace(e.emp_code, '^EMP-0(\\d{3})$', 'EMP-\\1') AS "Member_ID",
                   e.full_name AS "Full_name",
                   d.name AS "Department",
                   e.position_title AS "Role_title",
                   cl.level_code AS "Level",
                   regexp_replace(mgr.emp_code, '^EMP-0(\\d{3})$', 'EMP-\\1') AS "Line_manager_id",
                   s.name AS "Employment_status",
                   e.employment_type AS "Employment",
                   e.std_hours_week AS "Std_hours_week",
                   e.join_date AS "Join_date"
            FROM core.employee e
            JOIN core.department d ON d.department_id = e.department_id
            JOIN core.career_level cl ON cl.career_level_id = e.career_level_id
            JOIN core.employment_status s ON s.employment_status_id = e.employment_status_id
            LEFT JOIN core.employee mgr ON mgr.employee_id = e.line_manager_id
            WHERE e.employee_id IN (
              SELECT DISTINCT ra.employee_id FROM pmo.resource_allocation ra
              JOIN core.project p ON p.project_id = ra.project_id
              WHERE p.project_code ~ '^PRJ-'
              UNION
              SELECT pt.assignee_id FROM pmo.plan_task pt WHERE pt.assignee_id IS NOT NULL
              UNION
              SELECT p.pm_id FROM core.project p WHERE p.pm_id IS NOT NULL
                AND p.project_code ~ '^PRJ-'
            )
            ORDER BY e.emp_code
        """)),
        ("REF_Calendar_Weeks", fetch(cur, """
            SELECT 'W' || RANK() OVER (ORDER BY cw.week_start) AS "Week_ID",
                   cw.week_start AS "Week_start",
                   (cw.week_start + INTERVAL '4 days')::date AS "Week_end",
                   cw.working_days AS "Working_days",
                   cw.holiday_hours_ft AS "Holiday_hours_ft",
                   (SELECT STRING_AGG(ph.name || ' (' || ph.holiday_date || ')', '; ')
                    FROM core.public_holiday ph
                    WHERE ph.holiday_date BETWEEN cw.week_start
                      AND cw.week_start + INTERVAL '6 days'
                   ) AS "Note"
            FROM core.calendar_week cw
            ORDER BY cw.week_start
        """)),
        ("REF_KPI_Norms", fetch(cur, """
            SELECT mn.norm_code AS "Norm_ID",
                   mn.metric AS "Metric",
                   mn.formula AS "Formula",
                   MAX(CASE WHEN mnt.rag = 'Green' THEN mnt.rule_expr END) AS "Green",
                   MAX(CASE WHEN mnt.rag = 'Yellow' THEN mnt.rule_expr END) AS "Yellow",
                   MAX(CASE WHEN mnt.rag = 'Red' THEN mnt.rule_expr END) AS "Red",
                   mn.used_for AS "Used_for"
            FROM core.metric_norm mn
            LEFT JOIN core.metric_norm_threshold mnt
                ON mnt.metric_norm_id = mn.metric_norm_id
            GROUP BY mn.norm_code, mn.metric, mn.formula, mn.used_for
            ORDER BY mn.norm_code
        """)),
    ]
    write_xlsx(out_dir, name, sheets)


# ---------------------------------------------------------------------------
# DS03 — TA Hire Request & JD Generation
# ---------------------------------------------------------------------------
def export_ds03(cur, out_dir, name):
    sheets = [
        ("DS-01_Business_Context", fetch(cur, """
            SELECT bc.context_code AS context_id,
                   COALESCE(bc.project_name, p.name) AS project_name,
                   bc.business_unit, bc.roadmap_summary AS business_roadmap_summary,
                   bc.strategic_priority, bc.project_stage,
                   bc.requested_by, bc.request_date,
                   bc.team_size_current, bc.hiring_urgency,
                   bc.budget_approved, bc.additional_context
            FROM ta.business_context bc
            LEFT JOIN core.project p ON p.project_id = bc.project_id
            WHERE bc.context_code <= 'CTX-023'
            ORDER BY bc.context_code
        """)),
        ("DS-02_Headcount_Plan", fetch(cur, """
            SELECT hp.hc_plan_code AS hc_plan_id,
                   bc.context_code AS context_id,
                   hp.position, hp.seniority_level,
                   hp.headcount, hp.filled_count,
                   hp.salary_range, hp.approval_status,
                   hp.approved_by, hp.priority,
                   hp.target_start_date, hp.jd_status, hp.linked_jd_id
            FROM ta.headcount_plan hp
            JOIN ta.business_context bc ON bc.business_context_id = hp.context_id
            WHERE hp.hc_plan_code NOT IN ('HC-2025-Q2-031','HC-2025-Q2-032','HC-2025-Q2-033')
            ORDER BY hp.hc_plan_code
        """)),
        ("DS-03_JD_Template", fetch(cur, """
            SELECT jt.jd_code AS jd_id,
                   hp.hc_plan_code AS hc_plan_id,
                   bc.context_code AS context_id,
                   jt.position,
                   jt.jd_type,
                   jt.jd_version,
                   jt.jd_status AS status,
                   jt.last_updated,
                   jt.min_yoe, jt.max_yoe, jt.seniority_level,
                   jt.must_have_skills, jt.nice_to_have_skills,
                   jt.english_level_required, jt.work_mode,
                   CASE WHEN jt.salary_min_scaled IS NOT NULL AND jt.salary_max_scaled IS NOT NULL
                        THEN '$' || (jt.salary_min_scaled * 1000)::int
                             || '–$' || (jt.salary_max_scaled * 1000)::int || '/month'
                        ELSE NULL END AS salary_range,
                   jt.key_responsibilities, jt.jd_full_text
            FROM ta.jd_template jt
            LEFT JOIN ta.headcount_plan hp ON hp.headcount_plan_id = jt.hc_plan_id
            LEFT JOIN ta.business_context bc ON bc.business_context_id = hp.context_id
            WHERE jt.jd_code != 'JD-BE-SR-002'
            ORDER BY jt.jd_code
        """)),
        ("DS-04_Team_Skills_Matrix", fetch(cur, """
            SELECT member_id, member_role, team_name, seniority_level,
                   skill, proficiency_level, last_assessed
            FROM ta.team_skills_matrix
            ORDER BY member_id, skill
        """)),
        ("DS-05_Scorecard", fetch(cur, """
            SELECT sc.scorecard_code AS scorecard_id,
                   sc.position AS role,
                   sc.scorecard_version,
                   cr.interview_stage, cr.criteria, cr.weight,
                   cr.description, cr.passing_threshold, cr.sample_questions
            FROM ta.scorecard sc
            JOIN ta.scorecard_criterion cr ON cr.scorecard_id = sc.scorecard_id
            WHERE sc.scorecard_code IN (
              'SC-BE-SR-001','SC-AI-SR-001','SC-DE-SR-001',
              'SC-DS-SR-001','SC-DE-SR-003','SC-AIAG-SR-001','SC-BI-MID-001'
            )
            ORDER BY array_position(
              ARRAY['SC-BE-SR-001','SC-AI-SR-001','SC-DE-SR-001',
                    'SC-DS-SR-001','SC-DE-SR-003','SC-AIAG-SR-001','SC-BI-MID-001'],
              sc.scorecard_code
            ), cr.weight DESC
        """)),
        ("DS-06_Hire_Request", fetch(cur, """
            SELECT hr.request_code AS request_id,
                   bc.context_code AS context_id,
                   hp.hc_plan_code AS hc_plan_id,
                   hr.request_date, hr.requesting_manager,
                   hr.position_title, hr.urgency_level, hr.headcount_requested,
                   hr.business_justification, hr.team_skill_gap_summary,
                   hr.key_deliverables, hr.approval_status,
                   hr.approved_by, hr.approval_date,
                   hr.hr_owner, hr.target_jd_id,
                   hr.request_status, hr.notes
            FROM ta.hire_request hr
            LEFT JOIN ta.business_context bc ON bc.business_context_id = hr.context_id
            LEFT JOIN ta.headcount_plan hp ON hp.headcount_plan_id = hr.hc_plan_id
            ORDER BY hr.request_code
        """)),
        ("DS-07_Shortlist_CVs", fetch(cur, """
            SELECT cv.cv_code AS cv_id,
                   hr.request_code AS request_id,
                   hp.hc_plan_code AS hc_plan_id,
                   jt.jd_code AS jd_id,
                   c.candidate_code AS candidate_id,
                   cv.full_name, cv.current_title, cv.current_company,
                   cv.past_companies, cv.years_of_experience, cv.cv_skills,
                   cv.english_level, cv.salary_expectation,
                   cv.shortlisted_by, cv.shortlisted_date,
                   cv.cv_summary_by_ta, cv.agent_recommendation,
                   cv.agent_fit_score, cv.agent_fit_summary,
                   cv.agent_gap_summary, cv.agent_suggested_questions,
                   cv.agent_shortlist_rank
            FROM ta.shortlist_cv cv
            LEFT JOIN ta.hire_request hr ON hr.hire_request_id = cv.request_id
            LEFT JOIN ta.headcount_plan hp ON hp.headcount_plan_id = cv.hc_plan_id
            LEFT JOIN ta.jd_template jt ON jt.jd_template_id = cv.jd_id
            LEFT JOIN ta.candidate c ON c.candidate_id = cv.candidate_id
            ORDER BY cv.cv_code
        """)),
        ("DS-08_HM_Feedback_Tracker", fetch(cur, """
            SELECT fb.feedback_code AS feedback_id,
                   cv.cv_code AS cv_id,
                   hr.request_code AS request_id,
                   jt.jd_code AS jd_id,
                   fb.candidate_name, fb.position, fb.hiring_manager,
                   fb.shortlisted_datetime, fb.feedback_deadline_48h,
                   fb.sla_breach, fb.reminder_24h_sent,
                   fb.reminder_36h_sent, fb.escalation_48h_sent,
                   fb.feedback_status, fb.hm_decision,
                   fb.feedback_submitted_datetime, fb.hm_feedback_text
            FROM ta.hm_feedback_tracker fb
            LEFT JOIN ta.shortlist_cv cv ON cv.shortlist_cv_id = fb.cv_id
            LEFT JOIN ta.hire_request hr ON hr.hire_request_id = fb.request_id
            LEFT JOIN ta.jd_template jt ON jt.jd_template_id = fb.jd_id
            ORDER BY fb.feedback_code
        """)),
    ]
    write_xlsx(out_dir, name, sheets)


# ---------------------------------------------------------------------------
# DS04 — TA CV Screening
# ---------------------------------------------------------------------------
def export_ds04(cur, out_dir, name):
    sheets = [
        ("DS-06_Candidate_Database", fetch(cur, """
            SELECT c.candidate_code AS candidate_id,
                   c.full_name, c.email, c.phone,
                   c.location, c.applied_position,
                   c.current_title, c.current_company, c.past_companies,
                   c.years_of_experience, c.seniority_level, c.domain_experience,
                   c.employment_history, c.notable_projects,
                   CASE
                     WHEN c.salary_expectation_min_scaled IS NOT NULL
                          AND c.salary_expectation_max_scaled IS NOT NULL
                     THEN '$' || (c.salary_expectation_min_scaled * 1000)::int::text
                          || '–$' || (c.salary_expectation_max_scaled * 1000)::int::text
                     ELSE NULL
                   END AS salary_expectation,
                   c.cv_skills, c.english_level,
                   c.highest_education, c.education_major, c.certifications,
                   c.github_url, c.status, c.pipeline_stage, c.source,
                   c.received_cv_date, c.last_contact_date, c.result_release_date,
                   c.recruiter_owner, c.rejection_reason,
                   c.re_engagement_eligible, c.re_engagement_notes
            FROM ta.candidate c
            ORDER BY c.candidate_code
        """)),
        ("DS-06b_Candidate_Skills", fetch(cur, """
            SELECT c.candidate_code, s.skill_code, s.name AS skill_name,
                   sc.name AS category_name
            FROM ta.candidate_skill cs
            JOIN ta.candidate c ON c.candidate_id = cs.candidate_id
            JOIN core.skill s ON s.skill_id = cs.skill_id
            JOIN core.skill_category sc
                ON sc.skill_category_id = s.skill_category_id
            ORDER BY c.candidate_code, s.skill_code
        """)),
        ("DS-07_Screening_Criteria", fetch(cur, """
            SELECT sc.criteria_code AS criteria_id,
                   sc.jd_code AS jd_id, sc.position,
                   sc.must_have_skills, sc.nice_to_have_skills,
                   sc.tech_stack_preferred, sc.seniority_required,
                   sc.min_yoe, sc.max_yoe, sc.english_level_required,
                   sc.domain_preferred, sc.work_mode,
                   sc.salary_budget_max, sc.employment_type,
                   sc.weight_must_have_skills, sc.weight_yoe,
                   sc.weight_english, sc.weight_nice_to_have,
                   sc.scoring_note, sc.auto_flag_if_missing, sc.guardrail_notes
            FROM ta.screening_criteria sc
            ORDER BY sc.criteria_code
        """)),
        ("DS-07b_Screening_Criteria_Skills", fetch(cur, """
            SELECT sc.criteria_code, s.skill_code, s.name AS skill_name,
                   scs.skill_type
            FROM ta.screening_criteria_skill scs
            JOIN ta.screening_criteria sc
                ON sc.screening_criteria_id = scs.criteria_id
            JOIN core.skill s ON s.skill_id = scs.skill_id
            ORDER BY sc.criteria_code, scs.skill_type, s.skill_code
        """)),
        ("DS-08_Outreach_Template", fetch(cur, """
            SELECT template_code AS template_id,
                   channel, use_case, target_status, language,
                   template_content
            FROM ta.outreach_template
            ORDER BY template_code
        """)),
        ("DS-09_Job_Descriptions", fetch(cur, """
            SELECT jt.jd_code, jt.position, r.role_code,
                   jt.jd_version, jt.jd_status, jt.seniority_level,
                   jt.english_level_required, jt.work_mode,
                   jt.salary_min_scaled, jt.salary_max_scaled,
                   jt.must_have_skills, jt.nice_to_have_skills,
                   jt.key_responsibilities, jt.jd_full_text
            FROM ta.jd_template jt
            JOIN core.role r ON r.role_id = jt.role_id
            ORDER BY jt.jd_code
        """)),
        ("DS-10_Candidate_Fit_View", fetch(cur, """
            SELECT candidate_code, criteria_code,
                   must_have_overlap, must_have_total
            FROM ta.v_candidate_fit
            ORDER BY candidate_code, criteria_code
        """)),
    ]
    write_xlsx(out_dir, name, sheets)


# ---------------------------------------------------------------------------
# DS05 — ELC Employee Performance
# ---------------------------------------------------------------------------
def export_ds05(cur, out_dir, name):
    sheets = [
        ("DS00_Employee_Master", fetch(cur, """
            SELECT e.emp_code AS member_id,
                   e.position_title AS role_title,
                   d.name AS department,
                   cl.level_code AS level,
                   s.name AS employment_status,
                   e.join_date,
                   latest_pr.classification AS performance_tier,
                   latest_pr.total_point AS overall_score_latest
            FROM core.employee e
            JOIN core.department d ON d.department_id = e.department_id
            JOIN core.career_level cl ON cl.career_level_id = e.career_level_id
            JOIN core.employment_status s ON s.employment_status_id = e.employment_status_id
            LEFT JOIN LATERAL (
                SELECT pr.classification, pr.total_point
                FROM elc.performance_review pr
                WHERE pr.employee_id = e.employee_id
                ORDER BY pr.report_period DESC LIMIT 1
            ) latest_pr ON true
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$' AND e.emp_code <= 'EMP-100'
            ORDER BY e.emp_code
        """)),
        ("DS01_Resource_Allocation", fetch(cur, """
            SELECT e.emp_code AS member_id,
                   a.account_code AS account_id,
                   p.project_code AS project_id,
                   NULL::text AS assignment_type,
                   r.role_code AS role,
                   NULL::text AS report_to,
                   ra.allocation_pct,
                   NULL::text AS work_on_other,
                   NULL::text AS other_project_ids,
                   NULL::text AS notes
            FROM pmo.resource_allocation ra
            JOIN core.employee e ON e.employee_id = ra.employee_id
            JOIN core.project p ON p.project_id = ra.project_id
            JOIN core.account a ON a.account_id = p.account_id
            JOIN core.role r ON r.role_id = ra.role_id
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$' AND e.emp_code <= 'EMP-100'
            ORDER BY e.emp_code
        """)),
        ("DS02_Performance_by_Project", fetch(cur, """
            SELECT e.emp_code AS member_id,
                   NULL::text AS reviewer_id,
                   pr.report_period, pr.total_point,
                   pr.classification, pr.feedback_category, pr.review_frequency
            FROM elc.performance_review pr
            JOIN core.employee e ON e.employee_id = pr.employee_id
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$' AND e.emp_code <= 'EMP-100'
            ORDER BY e.emp_code, pr.report_period
        """)),
        ("DS03_Timesheet_Logwork", fetch(cur, """
            SELECT e.emp_code AS member_id, tm.report_period,
                   tm.work_days_in_month,
                   tm.days_probation, tm.days_official, tm.days_holiday_official,
                   tm.days_leave_approved, tm.days_late, tm.days_absent_unapproved,
                   tm.actual_work_days,
                   tm.ot_hours_weekday, tm.ot_hours_weekend, tm.ot_hours_holiday,
                   tm.total_ot_hours, tm.night_shift_hours
            FROM elc.timesheet_monthly tm
            JOIN core.employee e ON e.employee_id = tm.employee_id
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$' AND e.emp_code <= 'EMP-100'
            ORDER BY e.emp_code, tm.report_period
        """)),
        ("DS04_Violation_Attitude", fetch(cur, """
            SELECT v.violation_code AS violation_id,
                   e.emp_code AS member_id,
                   vt.category,
                   v.violation_type_code,
                   vt.violation_type_desc,
                   v.severity, v.consequence, v.status, v.incident_date,
                   v.reported_by, v.action_taken
            FROM elc.violation v
            JOIN core.employee e ON e.employee_id = v.employee_id
            JOIN elc.violation_type vt ON vt.violation_type_code = v.violation_type_code
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$' AND e.emp_code <= 'EMP-100'
            ORDER BY v.incident_date, e.emp_code
        """)),
        ("DS04b_ViolationType_Ref", fetch(cur, """
            SELECT vt.category, vt.violation_type_code,
                   vt.violation_type_desc, vt.typical_severity, vt.typical_consequence
            FROM elc.violation_type vt
            ORDER BY vt.violation_type_code
        """)),
        ("DS04c_Violation_Summary", fetch(cur, """
            SELECT emp_code AS member_id, total_violations,
                   critical_count, high_count, medium_count, low_count,
                   open_cases, risk_flag
            FROM elc.v_violation_summary
            WHERE emp_code ~ '^EMP-[0-9]{3}$' AND emp_code <= 'EMP-100'
            ORDER BY emp_code
        """)),
        ("DS05_Promotion_Intent", fetch(cur, """
            SELECT e.emp_code AS member_id,
                   cl_cur.level_code AS current_level,
                   cl_tgt.level_code AS target_level,
                   pi.readiness_score
            FROM elc.promotion_intent pi
            JOIN core.employee e ON e.employee_id = pi.employee_id
            JOIN core.career_level cl_cur ON cl_cur.career_level_id = pi.current_level_id
            JOIN core.career_level cl_tgt ON cl_tgt.career_level_id = pi.target_level_id
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$' AND e.emp_code <= 'EMP-100'
            ORDER BY e.emp_code
        """)),
        ("DS06_Salary_Band", fetch(cur, """
            SELECT e.emp_code AS member_id, sb.salary_band, sb.effective_date
            FROM elc.salary_band sb
            JOIN core.employee e ON e.employee_id = sb.employee_id
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$' AND e.emp_code <= 'EMP-100'
            ORDER BY e.emp_code, sb.effective_date
        """)),
        ("DS07_Performance_NORM", fetch(cur, """
            SELECT norm_code AS norm_id, category, rule_description, threshold,
                   classification_label, action_if_triggered, priority, applies_to
            FROM elc.performance_norm
            ORDER BY norm_code
        """)),
        ("DS08_Perf_Profile_Agg", fetch(cur, """
            SELECT vp.emp_code AS member_id,
                   vp.avg_score AS avg_score_t3_t4,
                   vp.classification_latest,
                   vp.ts_compliance_label AS ts_compliance_t4,
                   vp.total_ot_hours_latest AS total_ot_hours_t4,
                   vp.violation_risk_flag,
                   vp.open_violation_count,
                   vp.allocation_status,
                   pit.readiness_score,
                   sb.salary_band,
                   vp.perf_risk_note
            FROM elc.v_perf_profile vp
            LEFT JOIN elc.promotion_intent pit ON pit.employee_id = vp.employee_id
            LEFT JOIN elc.salary_band sb ON sb.employee_id = vp.employee_id
              AND sb.effective_date = (
                SELECT MAX(sb2.effective_date) FROM elc.salary_band sb2
                WHERE sb2.employee_id = vp.employee_id
              )
            WHERE vp.emp_code ~ '^EMP-[0-9]{3}$' AND vp.emp_code <= 'EMP-100'
            ORDER BY vp.emp_code
        """)),
        ("REF_Project_Master", fetch(cur, """
            SELECT a.account_code AS account_id,
                   a.name AS account_name,
                   p.project_code AS project_id,
                   p.name AS project_name
            FROM core.project p
            JOIN core.account a ON a.account_id = p.account_id
            WHERE p.project_code ~ '^(ACC|INT)-'
            ORDER BY a.account_code, p.project_code
        """)),
    ]
    write_xlsx(out_dir, name, sheets)


# ---------------------------------------------------------------------------
# DS06 — LnD Training Roadmap
# ---------------------------------------------------------------------------
def export_ds06(cur, out_dir, name):
    sheets = [
        ("DS01_Employee_Skill_Profile", fetch(cur, """
            SELECT
              e.emp_code AS "Employee_ID",
              COALESCE(e.position_title, r.name) AS "Position",
              COALESCE(STRING_AGG(DISTINCT s.name, '; ' ORDER BY s.name), '') AS "Skill",
              COALESCE(MAX(pl.prof_code), '') AS "Proficiency_Level",
              COALESCE((
                SELECT STRING_AGG(DISTINCT gs.name, ', ' ORDER BY gs.name)
                FROM lnd.employee_skill_gap g2
                JOIN core.skill gs ON gs.skill_id = g2.skill_id
                WHERE g2.employee_id = e.employee_id
              ), '') AS "Skill_Gap"
            FROM core.employee e
            JOIN core.role r ON r.role_id = e.role_id
            LEFT JOIN core.employee_skill es ON es.employee_id = e.employee_id
            LEFT JOIN core.skill s ON s.skill_id = es.skill_id
            LEFT JOIN core.proficiency_level pl
                ON pl.proficiency_level_id = es.proficiency_level_id
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$'
            GROUP BY e.emp_code, e.position_title, r.name, e.employee_id
            ORDER BY e.emp_code
        """)),
        ("DS02_Project_Roadmap", fetch(cur, """
            SELECT
              p.project_code AS "Project_ID",
              STRING_AGG(sq.skill_name, ', ' ORDER BY sq.skill_name) AS "Required_Skills",
              COALESCE(p.timeline, '') AS "Timeline"
            FROM core.project p
            JOIN (
              SELECT DISTINCT rs.project_id, s.name AS skill_name
              FROM lnd.project_required_skill rs
              JOIN core.skill s ON s.skill_id = rs.skill_id
            ) sq ON sq.project_id = p.project_id
            WHERE p.project_code ~ '^PRJ-[0-9]{3}$'
            GROUP BY p.project_code, p.timeline
            ORDER BY p.project_code
        """)),
        ("DS03_Training_Need_Survey", fetch(cur, """
            SELECT
              tns.survey_wave AS "Survey_ID",
              e.emp_code AS "Employee_ID",
              tns.training_topic AS "Training_Topic",
              tns.priority AS "Priority"
            FROM lnd.training_need_survey tns
            JOIN core.employee e ON e.employee_id = tns.employee_id
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$'
            ORDER BY tns.survey_wave, e.emp_code
        """)),
        ("DS04_Internal_Trainer_List", fetch(cur, """
            SELECT
              t.trainer_code AS "Trainer_ID",
              COALESCE(t.expertise, '') AS "Expertise",
              t.availability_hours_per_month AS "Availability_Hours_Per_Month"
            FROM core.trainer t
            ORDER BY t.trainer_code
        """)),
        ("DS05_BOD_Training_Goals", fetch(cur, """
            SELECT
              goal_code AS "Goal_ID",
              goal_description AS "Goal_Description",
              target_quarter AS "Target_Quarter"
            FROM lnd.bod_training_goal
            ORDER BY goal_code
        """)),
    ]
    write_xlsx(out_dir, name, sheets)


# ---------------------------------------------------------------------------
# DS07 — LnD Training Effectiveness
# ---------------------------------------------------------------------------
def export_ds07(cur, out_dir, name):
    sheets = [
        ("DS06_Course_Catalog", fetch(cur, """
            SELECT cc.course_code AS "Course_ID",
                   cc.course_name AS "Course_Name",
                   cc.topic_category AS "Topic_Category",
                   t.trainer_code AS "Trainer_ID",
                   cc.total_sessions AS "Total_Sessions",
                   cc.hours_per_session AS "Hours_Per_Session",
                   cc.total_hours AS "Total_Hours",
                   cc.pass_threshold_score AS "Pass_Threshold_Score",
                   cc.start_date AS "Start_Date",
                   cc.end_date AS "End_Date",
                   cc.status AS "Status"
            FROM lnd.course_catalog cc
            LEFT JOIN core.trainer t ON t.trainer_id = cc.trainer_id
            ORDER BY cc.course_code
        """)),
        ("DS07_Attendance_Log", fetch(cur, """
            SELECT cc.course_code AS "Course_ID",
                   cc.course_code || '_S' || al.session_no AS "Session_ID",
                   e.emp_code AS "Employee_ID",
                   al.attendance_status AS "Attendance_Status",
                   al.training_hours AS "Training_Hours"
            FROM lnd.attendance_log al
            JOIN lnd.course_catalog cc ON cc.course_id = al.course_id
            JOIN core.employee e ON e.employee_id = al.employee_id
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$'
            ORDER BY cc.course_code, al.session_no, e.emp_code
        """)),
        ("DS08_Assessment_Score", fetch(cur, """
            SELECT cc.course_code AS "Course_ID",
                   e.emp_code AS "Employee_ID",
                   asm.score_0_to_10 AS "Score_0_to_10",
                   asm.pass_status AS "Pass_Status",
                   asm.generalized_feedback AS "Generalized_Feedback"
            FROM lnd.assessment_score asm
            JOIN lnd.course_catalog cc ON cc.course_id = asm.course_id
            JOIN core.employee e ON e.employee_id = asm.employee_id
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$'
            ORDER BY cc.course_code, e.emp_code
        """)),
        ("DS09_Feedback_Survey", fetch(cur, """
            SELECT cc.course_code AS "Course_ID",
                   e.emp_code AS "Employee_ID",
                   fs.trainer_rating AS "Trainer_Rating_1_to_5",
                   fs.content_rating AS "Content_Rating_1_to_5",
                   fs.comment AS "Comment"
            FROM lnd.feedback_survey fs
            JOIN lnd.course_catalog cc ON cc.course_id = fs.course_id
            JOIN core.employee e ON e.employee_id = fs.employee_id
            WHERE e.emp_code ~ '^EMP-[0-9]{3}$'
            ORDER BY cc.course_code, e.emp_code
        """)),
        ("DS10_Training_Cost_ROI", fetch(cur, """
            SELECT cc.course_code AS "Course_ID",
                   tc.cost_per_session_scaled AS "Cost_Per_Session_Scaled",
                   cc.total_sessions AS "Total_Sessions",
                   tc.total_cost_scaled AS "Total_Cost_Scaled",
                   ce.trainee_count AS "Trainee_Count",
                   ce.completion_rate AS "Completion_Rate",
                   ce.avg_score AS "Avg_Score",
                   ce.pass_rate AS "Pass_Rate",
                   tc.post_training_perf_delta AS "Post_Training_Perf_Delta",
                   NULL::text AS "Notes"
            FROM lnd.training_cost tc
            JOIN lnd.course_catalog cc ON cc.course_id = tc.course_id
            LEFT JOIN lnd.v_course_effectiveness ce ON ce.course_id = tc.course_id
            ORDER BY cc.course_code
        """)),
        ("DS11_LnD_Training_NORM", fetch(cur, """
            SELECT rule_code AS "Rule_ID",
                   category AS "Category",
                   rule_description AS "Rule_Description",
                   threshold AS "Threshold",
                   action_if_triggered AS "Action_If_Triggered",
                   priority AS "Priority"
            FROM lnd.training_norm
            ORDER BY rule_code
        """)),
        ("DS12_Report_Template_Structure", fetch(cur, """
            SELECT section_code AS "Section_ID",
                   section_name AS "Section_Name",
                   content_description AS "Content_Description",
                   data_source AS "Data_Source",
                   is_required AS "Required"
            FROM lnd.report_template_section
            ORDER BY section_code
        """)),
    ]
    write_xlsx(out_dir, name, sheets)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
DATASETS = [
    ("01_pmo_project_plan_review",    export_ds01),
    ("02_pmo_timesheet_monitoring",   export_ds02),
    ("03_ta_hire_request_jd_generation", export_ds03),
    ("04_ta_cv_screening",            export_ds04),
    ("05_elc_employee_performance",   export_ds05),
    ("06_lnd_training_roadmap",       export_ds06),
    ("07_lnd_training_effectiveness", export_ds07),
]


def main():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    out_dir = DATASETS_DIR / "output"
    for folder, fn in DATASETS:
        print(f"\n[{folder}]")
        fn(cur, out_dir, folder)  # fn writes write_xlsx(out_dir, folder, sheets)

    conn.close()
    print(f"\nDone — 7 xlsx files written to datasets/output/")


if __name__ == "__main__":
    main()
