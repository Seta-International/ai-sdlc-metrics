#!/usr/bin/env python3
"""Build docs/exels/project-plan.xlsx from the Markdown sheet sources.

One-shot generator: reads each Markdown file in docs/exels/, parses its blocks
(title, headings, paragraphs, bullets, tables, code fences), and writes each to
a corresponding sheet in the workbook with basic styling.

Run:  python3 scripts/build-project-plan-xlsx.py
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
EXELS = ROOT / "docs" / "exels"
OUTPUT = EXELS / "project-plan.xlsx"

SHEETS: list[tuple[str, str]] = [
    ("00-Cover", "00-Cover.md"),
    ("00-TOC", "00-TOC.md"),
    ("01-Overview", "01-Overview.md"),
    ("02-Contract", "02-Contract.md"),
    ("03-Scope", "03-Scope.md"),
    ("04-Timeline", "04-Timeline.md"),
    ("05-Approach", "05-Approach.md"),
    ("06-Resources", "06-Resources.md"),
    ("07-DCA", "07-DCA.md"),
    ("08-Risks-Issues", "08-Risks-Issues.md"),
    ("09-ExecSupport", "09-ExecSupport.md"),
    ("Track-Planner", "Track-Planner.md"),
    ("Track-CoreBackend", "Track-CoreBackend.md"),
    ("Track-CoreFrontend", "Track-CoreFrontend.md"),
    ("Track-CoreAIAgent", "Track-CoreAIAgent.md"),
]

# --- Styling ---------------------------------------------------------------

PALETTE = {
    "title": "1f2937",
    "h2": "1f4e8a",
    "h3": "334155",
    "header_fill": "1f4e8a",
    "header_text": "ffffff",
    "band_fill": "f3f4f6",
    "border": "d1d5db",
    "blockquote_fill": "fef9c3",
    "code_fill": "f1f5f9",
    # Status legend
    "status_green": "16a34a",
    "status_yellow": "eab308",
    "status_red": "dc2626",
    "status_grey": "6b7280",
    "status_blue": "2563eb",
}

BORDER_THIN = Border(
    left=Side(style="thin", color=PALETTE["border"]),
    right=Side(style="thin", color=PALETTE["border"]),
    top=Side(style="thin", color=PALETTE["border"]),
    bottom=Side(style="thin", color=PALETTE["border"]),
)


def status_fill(value: str) -> PatternFill | None:
    lower = (value or "").strip().lower()
    if lower in {"done", "on-track", "on track", "completed"}:
        return PatternFill("solid", fgColor=PALETTE["status_green"])
    if lower in {"at risk", "at-risk", "slight delay"}:
        return PatternFill("solid", fgColor=PALETTE["status_yellow"])
    if lower in {"blocked", "critical"}:
        return PatternFill("solid", fgColor=PALETTE["status_red"])
    if lower in {"planned"}:
        return PatternFill("solid", fgColor=PALETTE["status_grey"])
    if lower in {"in progress"}:
        return PatternFill("solid", fgColor=PALETTE["status_blue"])
    return None


# --- Markdown parsing ------------------------------------------------------

INLINE_PATTERNS = [
    (re.compile(r"\*\*(.+?)\*\*"), r"\1"),  # strip bold markers but keep text
    (re.compile(r"__(.+?)__"), r"\1"),
    (re.compile(r"(?<!\*)\*(?!\*)(.+?)\*"), r"\1"),  # italics
    (re.compile(r"_(?!_)(.+?)_"), r"\1"),
    (re.compile(r"`([^`]+)`"), r"\1"),  # backticks
    (re.compile(r"\[([^\]]+)\]\([^)]+\)"), r"\1"),  # [text](url) -> text
]


def clean_inline(text: str) -> str:
    out = text
    for pat, repl in INLINE_PATTERNS:
        out = pat.sub(repl, out)
    return out.strip()


@dataclass
class Block:
    kind: str  # 'title' | 'h2' | 'h3' | 'para' | 'bullets' | 'table' | 'code' | 'blockquote' | 'hr'
    payload: object


def parse_md(path: Path) -> list[Block]:
    lines = path.read_text().splitlines()
    blocks: list[Block] = []
    i = 0
    while i < len(lines):
        line = lines[i].rstrip()
        stripped = line.strip()

        # Blank line
        if not stripped:
            i += 1
            continue

        # Horizontal rule
        if re.match(r"^-{3,}$|^\*{3,}$", stripped):
            blocks.append(Block("hr", None))
            i += 1
            continue

        # Headings
        if stripped.startswith("# "):
            blocks.append(Block("title", clean_inline(stripped[2:])))
            i += 1
            continue
        if stripped.startswith("## "):
            blocks.append(Block("h2", clean_inline(stripped[3:])))
            i += 1
            continue
        if stripped.startswith("### "):
            blocks.append(Block("h3", clean_inline(stripped[4:])))
            i += 1
            continue
        if stripped.startswith("#### "):
            blocks.append(Block("h3", clean_inline(stripped[5:])))
            i += 1
            continue

        # Code fence
        if stripped.startswith("```"):
            j = i + 1
            code_lines: list[str] = []
            while j < len(lines) and not lines[j].strip().startswith("```"):
                code_lines.append(lines[j])
                j += 1
            blocks.append(Block("code", "\n".join(code_lines)))
            i = j + 1
            continue

        # Blockquote
        if stripped.startswith("> "):
            buf: list[str] = []
            while i < len(lines) and lines[i].lstrip().startswith("> "):
                buf.append(lines[i].lstrip()[2:])
                i += 1
            blocks.append(Block("blockquote", clean_inline(" ".join(buf))))
            continue

        # Table
        if line.startswith("|"):
            table_rows: list[list[str]] = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                row = [clean_inline(c) for c in lines[i].strip().strip("|").split("|")]
                # Skip the alignment separator row
                if all(re.match(r"^:?-+:?$", c.strip()) for c in row):
                    i += 1
                    continue
                table_rows.append(row)
                i += 1
            if table_rows:
                blocks.append(Block("table", table_rows))
            continue

        # Bullet list
        if re.match(r"^[-*] ", stripped) or re.match(r"^\d+\. ", stripped):
            items: list[str] = []
            while i < len(lines):
                cur = lines[i].rstrip()
                if not cur.strip():
                    i += 1
                    break
                m_bul = re.match(r"^\s*[-*] (.*)", cur)
                m_num = re.match(r"^\s*\d+\. (.*)", cur)
                if m_bul:
                    items.append(clean_inline(m_bul.group(1)))
                    i += 1
                elif m_num:
                    items.append(clean_inline(m_num.group(1)))
                    i += 1
                elif cur.startswith("  ") and items:
                    items[-1] += " " + clean_inline(cur.strip())
                    i += 1
                else:
                    break
            blocks.append(Block("bullets", items))
            continue

        # Paragraph (merge consecutive non-blank, non-special lines)
        para_lines = [stripped]
        i += 1
        while i < len(lines):
            nxt = lines[i].rstrip()
            if not nxt.strip():
                break
            if nxt.strip().startswith("#") or nxt.startswith("|") or nxt.strip().startswith("```") or nxt.strip().startswith("> "):
                break
            if re.match(r"^[-*] |^\d+\. ", nxt.strip()):
                break
            para_lines.append(nxt.strip())
            i += 1
        blocks.append(Block("para", clean_inline(" ".join(para_lines))))

    return blocks


# --- Rendering -------------------------------------------------------------


def set_col_widths(ws: Worksheet, col_count: int, text_widths: list[int] | None = None) -> None:
    if text_widths is None:
        text_widths = [max(18, 180 // max(col_count, 1)) for _ in range(col_count)]
    for c in range(1, col_count + 1):
        w = text_widths[c - 1] if c - 1 < len(text_widths) else 30
        ws.column_dimensions[get_column_letter(c)].width = max(12, min(60, w))


def render_block(ws: Worksheet, row: int, block: Block) -> int:
    if block.kind == "title":
        cell = ws.cell(row=row, column=1, value=str(block.payload))
        cell.font = Font(size=18, bold=True, color=PALETTE["title"])
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
        return row + 2

    if block.kind == "h2":
        cell = ws.cell(row=row, column=1, value=str(block.payload))
        cell.font = Font(size=14, bold=True, color=PALETTE["h2"])
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22
        return row + 1

    if block.kind == "h3":
        cell = ws.cell(row=row, column=1, value=str(block.payload))
        cell.font = Font(size=12, bold=True, color=PALETTE["h3"])
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 20
        return row + 1

    if block.kind == "para":
        text = str(block.payload)
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        # Approximate height
        line_count = max(1, (len(text) // 110) + text.count("\n") + 1)
        ws.row_dimensions[row].height = min(180, max(18, line_count * 16))
        return row + 2

    if block.kind == "blockquote":
        text = str(block.payload)
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = PatternFill("solid", fgColor=PALETTE["blockquote_fill"])
        cell.font = Font(italic=True, color="713f12")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        line_count = max(1, (len(text) // 110) + 1)
        ws.row_dimensions[row].height = min(120, max(22, line_count * 16))
        return row + 2

    if block.kind == "bullets":
        items: list[str] = block.payload  # type: ignore[assignment]
        for item in items:
            cell = ws.cell(row=row, column=1, value=f"•  {item}")
            cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            line_count = max(1, (len(item) // 105) + 1)
            ws.row_dimensions[row].height = min(120, max(18, line_count * 16))
            row += 1
        return row + 1

    if block.kind == "code":
        text = str(block.payload)
        # Write as a single merged cell with monospace font
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(name="Menlo", size=10)
        cell.fill = PatternFill("solid", fgColor=PALETTE["code_fill"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        lines = text.count("\n") + 1
        ws.row_dimensions[row].height = min(400, max(30, lines * 14))
        return row + 2

    if block.kind == "table":
        rows: list[list[str]] = block.payload  # type: ignore[assignment]
        if not rows:
            return row
        header = rows[0]
        cols = len(header)
        # Find Status column index if present
        status_col_idx: int | None = None
        for idx, h in enumerate(header):
            if h.strip().lower() == "status":
                status_col_idx = idx
                break
        # Write header
        for c, h in enumerate(header, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = Font(bold=True, color=PALETTE["header_text"])
            cell.fill = PatternFill("solid", fgColor=PALETTE["header_fill"])
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            cell.border = BORDER_THIN
        ws.row_dimensions[row].height = 28
        # Freeze cannot be per-table; handled globally later if needed
        row += 1
        # Body
        for bi, body in enumerate(rows[1:]):
            padded = body + [""] * (cols - len(body))
            max_len = 0
            for c, val in enumerate(padded[:cols], start=1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                cell.border = BORDER_THIN
                if bi % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor=PALETTE["band_fill"])
                if status_col_idx is not None and c - 1 == status_col_idx:
                    fill = status_fill(val)
                    if fill is not None:
                        cell.fill = fill
                        cell.font = Font(bold=True, color="ffffff")
                max_len = max(max_len, len(val))
            approx_lines = max(1, (max_len // 50) + 1)
            ws.row_dimensions[row].height = min(120, max(18, approx_lines * 16))
            row += 1
        # Set reasonable column widths for this sheet if not yet set
        current = [ws.column_dimensions[get_column_letter(c + 1)].width or 0 for c in range(cols)]
        for c in range(cols):
            col_letter = get_column_letter(c + 1)
            existing = ws.column_dimensions[col_letter].width or 0
            # Width proxy: look at longest cell in this column across this table
            w = max(
                [len(str(r[c])) for r in rows if c < len(r)] + [8]
            )
            width = min(60, max(existing, min(w + 2, 50)))
            ws.column_dimensions[col_letter].width = width
        return row + 1

    if block.kind == "hr":
        cell = ws.cell(row=row, column=1, value="─" * 80)
        cell.font = Font(color=PALETTE["border"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        return row + 2

    return row


def build_sheet(wb: Workbook, name: str, md_path: Path, is_first: bool) -> None:
    blocks = parse_md(md_path)
    if is_first:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)

    # Default column widths
    for c in range(1, 12):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # Sheet-level tweaks: freeze top row for table-heavy sheets
    if name in {"03-Scope", "04-Timeline", "06-Resources", "08-Risks-Issues", "Track-Planner", "Track-CoreBackend", "Track-CoreFrontend", "Track-CoreAIAgent"}:
        ws.freeze_panes = "A2"

    row = 1
    for block in blocks:
        row = render_block(ws, row, block)

    # Wider column A for narrative sheets
    if name in {"00-Cover", "01-Overview", "05-Approach", "09-ExecSupport"}:
        ws.column_dimensions["A"].width = 40

    # Hyperlinks in 00-TOC (third column = "Sheet") — replace text with a sheet-name hyperlink.
    if name == "00-TOC":
        # Walk rows and if the "Sheet" column has a bare name like `01-Overview`, set a hyperlink.
        for r in range(1, ws.max_row + 1):
            for c in range(1, 5):
                v = ws.cell(r, c).value
                if isinstance(v, str) and re.match(r"^[0-9A-Za-z_-]+$", v) and any(v == s[0] for s in SHEETS):
                    cell = ws.cell(r, c)
                    cell.hyperlink = f"#'{v}'!A1"
                    cell.font = Font(color="1f4e8a", underline="single", bold=True)


def main() -> None:
    wb = Workbook()
    for idx, (name, filename) in enumerate(SHEETS):
        path = EXELS / filename
        if not path.exists():
            print(f"[skip] missing {path}")
            continue
        print(f"[build] {name}")
        build_sheet(wb, name, path, is_first=(idx == 0))

    # Sanity: remove the default "Sheet" if it's still there
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\nWrote {OUTPUT} with {len(wb.sheetnames)} sheets:")
    for s in wb.sheetnames:
        print(f"  - {s}")


if __name__ == "__main__":
    main()
