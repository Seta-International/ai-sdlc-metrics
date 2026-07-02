"""Pure workbook filling: template + fetched rows -> openpyxl Workbook."""
import re
from datetime import datetime
from pathlib import Path
import openpyxl
from exporter.charts import add_charts

SHEET3_METRIC_COLS = {
    "D": "ai_users_weekly_avg", "F": "ai_prs", "G": "total_prs",
    "H": "agent_tasks", "I": "total_tasks", "J": "lead_time_h", "K": "deploys",
    "L": "weeks", "M": "incidents", "N": "mttr_h", "Q": "rework_prs",
    "R": "ai_prs_reviewed", "T": "security_alerts", "U": "agent_prs_merged",
    "V": "agent_prs_human_fixed", "W": "agent_prs_autonomous", "X": "agent_cycle_h",
}
SHEET3_MANUAL_COLS = {"E": "total_engineers", "O": "cost_baseline",
                      "P": "cost_actual", "S": "coverage_ai"}
SHEET4_FIELDS = [
    "g1_agents_md", "g2_ai_policy", "g3_required_review", "g4_eval_suite",
    "g5_shared_library", "g6_security_controls", "g7_traceability",
    "g8_model_governance", "a2_dashboard", "a4_near_universal",
    "b4_dora_improving", "b5_cost_multi_wf", "b6_business_outcomes",
    "b7_top_quartile", "b8_client_reporting", "c3_scan_ci", "c4_ai_vs_nonai",
    "c5_evals", "c6_sast_pii_required", "c7_defect_zero", "c8_evals_in_ci",
    "c9_prompt_leak_pii", "d3_defined_class", "d4_cycle_measured",
    "d5_multi_agent", "evidence_a", "evidence_b", "evidence_c", "evidence_d",
    "evidence_e", "improvement_action",
]
SPRINT_SHEET_COLS = [
    "period_key", "period_start", "ai_users_weekly_avg", "ai_prs", "total_prs",
    "ai_pr_pct", "agent_tasks", "ai_tasks", "total_tasks", "agent_task_pct",
    "lead_time_h", "deploys", "weeks", "deploys_per_week", "incidents",
    "cfr_pct", "mttr_h", "rework_prs", "rework_pct", "ai_prs_reviewed",
    "ai_pr_review_pct", "security_alerts", "agent_prs_total",
    "agent_prs_merged", "agent_prs_human_fixed", "agent_prs_autonomous",
    "agent_completion_pct", "human_intervention_pct", "autonomy_pct",
    "agent_cycle_h", "sprint_committed", "sprint_completed", "predictability_pct",
]


def parse_sprint_range(spec: str | None) -> tuple[int, int] | None:
    if not spec:
        return None
    m = re.fullmatch(r"S(\d+)(?::S(\d+))?", spec)
    if not m:
        raise ValueError(f"sprints must look like 'S3' or 'S1:S6', got {spec!r}")
    lo = int(m.group(1))
    hi = int(m.group(2)) if m.group(2) else lo
    if lo > hi:
        raise ValueError(f"empty sprint range {spec!r}")
    return lo, hi


def sprint_in_range(period_key: str, rng: tuple[int, int] | None) -> bool:
    if rng is None:
        return True
    index = int(period_key[1:])
    return rng[0] <= index <= rng[1]


def _month_iter(start, end):
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        yield f"{y}-{m:02d}"
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)


def months_overlapped(sprint_rows: list[dict]) -> list[str]:
    months: set[str] = set()
    for row in sprint_rows:
        months.update(_month_iter(row["period_start"], row["period_end"]))
    return sorted(months)


def quarters_of(months: list[str]) -> list[str]:
    quarters = {f"{m[:4]}-Q{(int(m[5:7]) + 2) // 3}" for m in months}
    return sorted(quarters)


def _num(value):
    return float(value) if value is not None else None


def fill_workbook(template_path: Path, projects: list[str],
                  sprint_rows: list[dict], month_rows: list[dict],
                  manual: dict) -> openpyxl.Workbook:
    wb = openpyxl.load_workbook(template_path)
    ids = {name: f"P{i + 1:02d}" for i, name in enumerate(sorted(projects))}

    ws = wb["2. Projects"]
    for i, name in enumerate(sorted(projects)):
        ws[f"A{3 + i}"], ws[f"B{3 + i}"] = ids[name], name

    ws = wb["3. Monthly"]
    for i, row in enumerate(sorted(month_rows,
                                   key=lambda r: (r["project"], r["period_key"]))):
        r = 4 + i
        ws[f"A{r}"] = ids[row["project"]]
        ws[f"B{r}"] = datetime.strptime(row["period_key"], "%Y-%m")
        for col, key in SHEET3_METRIC_COLS.items():
            ws[f"{col}{r}"] = _num(row.get(key))
        month_manual = manual.get((row["project"], row["period_key"]), {})
        for col, field in SHEET3_MANUAL_COLS.items():
            if field in month_manual:
                ws[f"{col}{r}"] = float(month_manual[field])

    ws = wb["4. Quarterly"]
    quarter_keys = sorted({(p, q) for (p, q) in manual if re.fullmatch(r"\d{4}-Q[1-4]", q)
                           and p in projects})
    for i, (project, quarter) in enumerate(quarter_keys):
        r = 4 + i
        ws[f"A{r}"], ws[f"B{r}"] = ids[project], quarter
        entries = manual[(project, quarter)]
        for j, field in enumerate(SHEET4_FIELDS):
            if field in entries:
                ws.cell(row=r, column=3 + j, value=entries[field])

    ws = wb.create_sheet("Sprint data")
    header = ["Project"] + [c.replace("_", " ").title() for c in SPRINT_SHEET_COLS]
    for j, title in enumerate(header, start=1):
        ws.cell(row=1, column=j, value=title)
    for i, row in enumerate(sorted(sprint_rows,
                                   key=lambda r: (r["project"], r["period_start"])), start=2):
        ws.cell(row=i, column=1, value=row["project"])
        for j, key in enumerate(SPRINT_SHEET_COLS, start=2):
            v = row.get(key)
            if key == "period_key":
                ws.cell(row=i, column=j, value=v)
            elif key == "period_start":
                ws.cell(row=i, column=j, value=str(v))
            else:
                ws.cell(row=i, column=j, value=_num(v))

    add_charts(wb)
    return wb
