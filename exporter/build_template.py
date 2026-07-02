#!/usr/bin/env python3
"""
Build the English maturity template from the Vietnamese workbook.

  python -m exporter.build_template            # writes docs/SETA_AI_SDLC_Maturity_EN.xlsx
  python -m exporter.build_template --report   # list untranslated Vietnamese cells

Procedure: extend TRANSLATIONS until --report prints nothing, then commit the
EN file. Formulas and numbers are never touched; strings absent from the map
are left as-is (English/neutral strings need no entry).
"""
import argparse
import re
from pathlib import Path
import openpyxl
from exporter.charts import add_charts

SRC = Path("docs/SETA_AI_SDLC_Maturity.xlsx")          # Vietnamese source (input)
DST = Path("docs/AI SDLC Maturity.xlsx")               # English deliverable (output)

_VN = re.compile("[àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡ"
                 "ùúụủũưừứựửữỳýỵỷỹđ]", re.IGNORECASE)

SHEET_RENAMES = {"1. Cách đo & HD": "1. Guide"}

# Vietnamese string literals embedded inside formulas (e.g. HYPERLINK display
# text). Applied by substring replacement within formula cells.
FORMULA_TRANSLATIONS = {
    "ĐỊNH TÍNH  (tick → 4. Quarterly)": "QUALITATIVE  (tick → 4. Quarterly)",
    "ĐỊNH LƯỢNG  (auto ← 3. Monthly / 5. Metrics)":
        "QUANTITATIVE  (auto ← 3. Monthly / 5. Metrics)",
}

# Seed map — known headers. Extend until --report is empty (translate every
# remaining cell of sheets 1, 6, 7, 8, 9, 10 the same way).
TRANSLATIONS = {
    # 2. Projects
    "DANH MỤC DỰ ÁN": "PROJECT DIRECTORY",
    "Tên dự án": "Project name", "Tool chính": "Main tool", "Client / Mảng": "Client / Area",
    # 3. Monthly
    "NHẬP SỐ THEO THÁNG — chỉ gõ số thô (ô vàng). Metric % xem ở «5. Metrics».":
        "MONTHLY RAW INPUT — enter raw numbers only (yellow cells). Percentages are in «5. Metrics».",
    "Tháng (đầu tháng)": "Month (first day)", "Quý": "Quarter",
    "KS dùng AI/tuần": "Engineers using AI/week", "Tổng KS": "Total engineers",
    "PR gắn AI": "AI-labeled PRs", "Tổng PR": "Total PRs",
    "Task giao agent": "Tasks assigned to agents", "Tổng task": "Total tasks",
    "Số deploy": "Deploy count", "Số tuần": "Weeks", "Deploy lỗi": "Failed deploys",
    "Cost baseline/đv": "Cost baseline/unit", "Cost actual/đv": "Cost actual/unit",
    "PR rework ≤14d": "PR rework ≤14d", "PR-AI có review": "AI PRs reviewed",
    "Vulns/secrets chặn": "Vulns/secrets blocked",
    "Agent xong đạt": "Agent tasks done OK", "Agent cần sửa": "Agent tasks needing fixes",
    "Agent end-to-end": "Agent end-to-end", "Agent cycle (h)": "Agent cycle (h)",
    # 4. Quarterly
    "CHẤM ĐIỂM THEO QUÝ — chỉ tick flag PHÁN ĐOÁN (ô cam). Flag 'đã-có-số' tự suy.":
        "QUARTERLY SCORING — tick JUDGMENT flags only (orange cells). 'Measured' flags are derived.",
    "E. GOVERNANCE — checklist 8 mục": "E. GOVERNANCE — 8-item checklist",
    "G3 Review bắt buộc": "G3 Required review",
    "b4 DORA cải thiện": "b4 DORA improving", "b5 cost đạt nhiều WF": "b5 cost target, many workflows",
    "b8 báo cáo client": "b8 client reporting", "c3 scan cơ bản CI": "c3 basic CI scanning",
    "c4 so AI vs non-AI": "c4 AI vs non-AI compared", "c5 evals có": "c5 evals exist",
    "c6 SAST/PII bắt buộc": "c6 SAST/PII required", "c7 defect ~0": "c7 defects ~0",
    "d3 defined class": "d3 defined task class", "d4 cycle time đo": "d4 cycle time measured",
    # 10. Thresholds
    "NGƯỠNG QUY ĐỔI LEVEL (chỉnh được)": "LEVEL THRESHOLDS (adjustable)",
    "Tham số": "Parameter", "Giá trị": "Value", "Ý nghĩa": "Meaning",
    "Adoption L3: %PR có AI ≥ 30%": "Adoption L3: AI PR % ≥ 30%",
    "Adoption L4: %PR có AI > 50%": "Adoption L4: AI PR % > 50%",
    "Adoption L2: tỷ lệ người dùng AI ≥ 50%": "Adoption L2: AI user rate ≥ 50%",
    # Shared short labels (appear on several sheets)
    "Tháng": "Month", "%PR có AI": "AI PR %", "Chọn quý:": "Select quarter:",
    "Dự án": "Project", "Chiều": "Dimension",
    # 5. Metrics
    "METRIC (tự tính từ «3. Monthly») — không gõ tay.":
        "METRICS (auto-computed from «3. Monthly») — do not type.",
    "Deploy/tuần": "Deploys/week", "Cost cải thiện %": "Cost improvement %",
    # 6. Levels
    "QUY ĐỔI LEVEL (auto) — tra theo key ProjectID+Quý. Không gõ tay.":
        "LEVEL MAPPING (auto) — looked up by key ProjectID+Quarter. Do not type.",
    # 8. Dashboard-Project
    "HỒ SƠ DỰ ÁN": "PROJECT PROFILE", "Chọn dự án:": "Select project:",
    "TREND THEO THÁNG (dự án đã chọn)": "MONTHLY TREND (selected project)",
    # 9. Dashboard-Portfolio
    "PORTFOLIO — toàn bộ dự án theo quý": "PORTFOLIO — all projects by quarter",
    # 7. SelfAssessment
    "TỰ ĐÁNH GIÁ BASELINE — kickoff + soát lại mỗi quý":
        "BASELINE SELF-ASSESSMENT — kickoff + reviewed each quarter",
    "Mỗi 'No' là một hành động để lập baseline trước khi đo.":
        "Each 'No' is an action to establish a baseline before measuring.",
    "Chủ đề": "Topic", "Câu hỏi tự đánh giá": "Self-assessment question",
    "Ghi chú": "Notes", "Câu hỏi": "Question",
    "'Tốt' trông như thế nào": "What 'good' looks like",
    "Có AI policy văn bản (tool, dữ liệu, model được/cấm)? Mọi người tuân thủ? Prototype tách production?":
        "Written AI policy (allowed/forbidden tools, data, models)? Everyone complies? Prototype separated from production?",
    "Có AGENTS.md (stack, convention, hard rules) versioned & cập nhật khi AI sai? AI nối docs/codebase nội bộ?":
        "AGENTS.md (stack, conventions, hard rules) versioned & updated when AI errs? AI connected to internal docs/codebase?",
    "Viết test/eval TRƯỚC khi sinh code task quan trọng? Có eval suite pass/fail cho agent? Test chạy tự động CI?":
        "Write tests/evals BEFORE generating code for important tasks? Pass/fail eval suite for agents? Tests run automatically in CI?",
    "Mọi dòng code AI có người review trước ship? Reviewer biết lỗi đặc thù AI? Có cổng phê duyệt bắt buộc?":
        "Every line of AI code human-reviewed before ship? Reviewers know AI-specific failures? Mandatory approval gate?",
    "Có quét secret/PII & chống prompt-leak? Truy được artifact AI-sinh + log mọi agent run?":
        "Scanning for secrets/PII & prompt-leak protection? AI-generated artifacts traceable + every agent run logged?",
    "Đã ghi baseline (DORA, time, cost) trước AI? Theo dõi chất lượng song song năng suất? Gắn business outcomes?":
        "Baseline (DORA, time, cost) recorded before AI? Quality tracked alongside productivity? Tied to business outcomes?",
    "AI giúp sinh user story, tìm edge case, dựng prototype nhanh?":
        "AI helps generate user stories, find edge cases, build prototypes quickly?",
    "Từ mô tả → prototype vài phút; người set scope.":
        "From description → prototype in minutes; humans set scope.",
    "Kiến trúc do người, AI implement theo doc?": "Architecture by humans, AI implements per docs?",
    "Người sở hữu trade-off; AI scaffold theo doc.": "Humans own trade-offs; AI scaffolds per docs.",
    "AI sinh code trong spec và luôn được review?": "AI generates code within spec and always reviewed?",
    "AI làm phần lặp; người làm 20% khó.": "AI does the repetitive part; humans do the hard 20%.",
    "AI sinh test/edge case, kiểm output + trajectory?":
        "AI generates tests/edge cases, checks output + trajectory?",
    "Eval truyền 'đúng' & chạy tự động CI.": "Evals encode 'correct' & run automatically in CI.",
    "AI review vòng đầu trước khi người xem?": "AI reviews the first pass before humans look?",
    "AI reviewer đầu tiên; người quyết design.": "AI is the first reviewer; humans decide design.",
    "AI giúp hiểu & refactor legacy an toàn?": "AI helps understand & refactor legacy safely?",
    "Code 'quá rủi ro' nay refactor trong kiểm soát.":
        "Code once 'too risky' now refactored under control.",
    # 1. Guide
    "CÁCH ĐO & HƯỚNG DẪN": "HOW TO MEASURE & GUIDE",
    "QUY ƯỚC MÀU:": "COLOR CONVENTIONS:",
    "Ô VÀNG — nhập tay (số thô)": "YELLOW CELL — manual entry (raw numbers)",
    "Ô XÁM — công thức tự tính": "GRAY CELL — auto-computed formula",
    "Ô CAM — flag Yes/No phán đoán": "ORANGE CELL — Yes/No judgment flag",
    "Đi tới:": "Go to:",
    "ĐỊNH NGHĨA LEVEL THEO 2 PHƯƠNG PHÁP": "LEVEL DEFINITIONS BY 2 METHODS",
    "A. ADOPTION  —  AI có thực sự được dùng?": "A. ADOPTION  —  Is AI actually being used?",
    "Dùng lác đác, chưa rõ ai dùng": "Sporadic use, unclear who uses it",
    "— (a1 tự suy: chưa có số usage)": "— (a1 auto-derived: no usage numbers yet)",
    "a1 auto = có số usage  &  tỷ lệ dùng ≥ 50%": "a1 auto = usage numbers exist  &  usage rate ≥ 50%",
    "a2 = Yes (có dashboard theo dõi)": "a2 = Yes (tracking dashboard exists)",
    "%PR có AI ≥ 30%": "AI PR % ≥ 30%",
    "— (a3 tự suy: đã đo %việc giao agent)": "— (a3 auto-derived: % work assigned to agents measured)",
    "%PR có AI > 50%": "AI PR % > 50%",
    "a4 = Yes (adoption gần phổ cập, có chủ đích)": "a4 = Yes (adoption near-universal, intentional)",
    "B. DELIVERY & VALUE  —  Giao hàng nhanh & ổn định hơn? (DORA)":
        "B. DELIVERY & VALUE  —  Faster & more stable delivery? (DORA)",
    "Chưa đo gì": "Nothing measured yet",
    "b1 auto = đã có Deploy/tuần": "b1 auto = Deploys/week exists",
    "b2 auto (đủ 4 DORA) & b3 auto (có cost baseline)":
        "b2 auto (all 4 DORA) & b3 auto (cost baseline exists)",
    "(kèm 4 DORA cải thiện — xem trend)": "(with 4 DORA improving — see trend)",
    "b7 + b8 = Yes (báo cáo client định kỳ)": "b7 + b8 = Yes (regular client reporting)",
    "DORA top-quartile & ổn định": "DORA top-quartile & stable",
    "C. QUALITY & SECURITY  ★ CỬA CHẶN  —  Output đúng, sạch, an toàn?":
        "C. QUALITY & SECURITY  ★ GATE  —  Output correct, clean, safe?",
    "Không đo, đánh giá bằng mắt": "Not measured, assessed by eye",
    "Có ≥1 trong: G3 / c3": "Has ≥1 of: G3 / c3",
    "c2 auto = đã có Rework%": "c2 auto = Rework% exists",
    "G3 + c3 (review bắt buộc, scan cơ bản)": "G3 + c3 (required review, basic scanning)",
    "c2 auto (Rework% track); %PR-AI có review": "c2 auto (Rework% tracked); AI PRs reviewed %",
    "c4 + c5 + c6 (evals; SAST/PII bắt buộc)": "c4 + c5 + c6 (evals; SAST/PII required)",
    "Rework% giảm; Coverage tăng": "Rework% down; Coverage up",
    "D. AGENT MATURITY  —  Agent tự chủ tới đâu?": "D. AGENT MATURITY  —  How autonomous are agents?",
    "— (d1 tự suy: chưa có task giao agent)": "— (d1 auto-derived: no tasks assigned to agents yet)",
    "— (d2 tự suy: chưa đo completion/interv)": "— (d2 auto-derived: completion/intervention not measured)",
    "d1 auto = có task giao agent": "d1 auto = tasks assigned to agents exist",
    "d2 auto = có Agent success% + Human interv%": "d2 auto = Agent success% + Human intervention% exist",
    "d3 + d4 (defined class + cycle time đo)": "d3 + d4 (defined class + cycle time measured)",
    "E. GOVERNANCE  ★ CỬA CHẶN  —  Dùng AI có kỷ luật? (checklist 8 mục)":
        "E. GOVERNANCE  ★ GATE  —  Disciplined AI use? (8-item checklist)",
    "0–1 mục Yes": "0–1 items Yes", "Điểm checklist = 0–1 / 8": "Checklist score = 0–1 / 8",
    "2 mục Yes": "2 items Yes", "Điểm = 2 / 8": "Score = 2 / 8",
    "Đủ 3 Core: G1 AGENTS.md · G2 policy · G3 review":
        "All 3 Core: G1 AGENTS.md · G2 policy · G3 review",
    "Điểm ≥ 3 (đủ Core)": "Score ≥ 3 (Core complete)",
    "Đủ G1–G5 (Core + Advanced)": "All G1–G5 (Core + Advanced)",
    "Điểm = 5 (mục 1–5)": "Score = 5 (items 1–5)",
    "Đủ cả 8 (thêm Security & Audit)": "All 8 (plus Security & Audit)",
    "Điểm = 8 / 8": "Score = 8 / 8",
    """CÁCH CHẤM — mỗi chiều (A–E) đặt ở LEVEL CAO NHẤT thỏa ĐỒNG THỜI điều kiện định tính + định lượng:
•  ĐỊNH TÍNH = flag Yes/No do auditor tick (MANUAL) → sheet «4. Quarterly» (ô cam).
•  ĐỊNH LƯỢNG = số đo tự tính từ số thô (AUTO) → nhập ở «3. Monthly» (ô vàng); metric ở «5. Metrics»; level ở «6. Levels».
OVERALL = MIN( E-Governance , C-Quality , TRUNG BÌNH làm tròn 5 chiều ).  → C và E là CỬA CHẶN.
MANUAL tối thiểu = đúng 2 chỗ: số thô («3. Monthly») + flag phán đoán («4. Quarterly»).
Các flag kiểu "đã-có-số-hay-chưa" (a1,a3,b1,b2,b3,c2,d1,d2) TỰ suy từ dữ liệu — không phải tick.
Mã a2/a4/b4/c3/G3… = tên cột ở «4. Quarterly». Ngưỡng số (30/50/60/20%) chỉnh ở «10. Thresholds».""":
        """SCORING — place each dimension (A–E) at the HIGHEST LEVEL that SIMULTANEOUSLY meets the qualitative + quantitative conditions:
•  QUALITATIVE = Yes/No flag ticked by the auditor (MANUAL) → sheet «4. Quarterly» (orange cells).
•  QUANTITATIVE = measures auto-computed from raw numbers (AUTO) → enter in «3. Monthly» (yellow cells); metrics in «5. Metrics»; levels in «6. Levels».
OVERALL = MIN( E-Governance , C-Quality , ROUNDED AVERAGE of the 5 dimensions ).  → C and E are GATES.
MANUAL minimum = exactly 2 places: raw numbers («3. Monthly») + judgment flags («4. Quarterly»).
The "already-measured-or-not" flags (a1,a3,b1,b2,b3,c2,d1,d2) are AUTO-derived from data — not ticked.
Codes a2/a4/b4/c3/G3… = column names in «4. Quarterly». Numeric thresholds (30/50/60/20%) are adjusted in «10. Thresholds».""",
}

# Sample data to clear (values only; formula columns C / AH stay).
_CLEAR = [("2. Projects", 3, 20, "ABCDE"),
          ("3. Monthly", 4, 60, "AB" + "DEFGHIJKLMNOPQRSTUVWX"),
          ("4. Quarterly", 4, 53, [c for c in
           ["A", "B"] + [chr(x) for x in range(ord("C"), ord("Z") + 1)]
           + ["AA", "AB", "AC", "AD", "AE", "AF", "AG"]])]


def _translate_cell(cell) -> None:
    v = cell.value
    if isinstance(v, str) and not v.startswith("="):
        if v in TRANSLATIONS:
            cell.value = TRANSLATIONS[v]
    elif isinstance(v, str) and v.startswith("="):
        for old, new in {**SHEET_RENAMES, **FORMULA_TRANSLATIONS}.items():
            if old in v:
                v = v.replace(old, new)
        cell.value = v


def build() -> None:
    wb = openpyxl.load_workbook(SRC)
    for old, new in SHEET_RENAMES.items():
        wb[old].title = new
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                _translate_cell(cell)
    for sheet, r0, r1, cols in _CLEAR:
        ws = wb[sheet]
        for r in range(r0, r1 + 1):
            for col in cols:
                ws[f"{col}{r}"].value = None
    add_charts(wb)
    wb.save(DST)
    print(f"wrote {DST}")


_LITERAL = re.compile(r'"((?:[^"]|"")*)"')  # quoted string literals inside a formula


def report() -> None:
    wb = openpyxl.load_workbook(DST)
    remaining = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                if v.startswith("="):
                    for s in _LITERAL.findall(v):
                        if _VN.search(s):
                            print(f"{ws.title}!{cell.coordinate} [formula]: {s[:80]}")
                            remaining += 1
                elif _VN.search(v):
                    print(f"{ws.title}!{cell.coordinate}: {v[:80]}")
                    remaining += 1
    print(f"{remaining} untranslated cell(s)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--report", action="store_true")
    args = parser.parse_args()
    report() if args.report else build()
