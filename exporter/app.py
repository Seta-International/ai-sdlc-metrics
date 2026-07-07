"""AI SDLC maturity workbook exporter — serves the filled English template.

  uvicorn exporter.app:app --host 0.0.0.0 --port 8000
"""
import io, json, html
import os
import openpyxl
from fastapi import FastAPI, Form, HTTPException, Response, UploadFile
from fastapi.responses import HTMLResponse
from collector.db import upsert_manual_input
from exporter.data import fetch_manual, fetch_period_rows, fetch_projects, fetch_auto_ai_users
from exporter.importer import parse_manual_inputs, diff_changes, usage_warnings
from exporter.template import build_workbook
from exporter.workbook import fill_workbook, parse_month_range, month_in_range

app = FastAPI(title="AI SDLC Maturity Exporter")

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


@app.get("/export.xlsx")
def export(project: str = "all", months: str | None = None) -> Response:
    db_url = os.environ["REPORTING_DB_URL"]
    try:
        rng = parse_month_range(months)
    except ValueError as e:
        raise HTTPException(422, str(e))

    known = fetch_projects(db_url)
    projects = known if project == "all" else [project]
    if project != "all" and project not in known:
        raise HTTPException(404, f"unknown project {project!r}")

    month_rows = [r for r in fetch_period_rows(db_url, projects, "month")
                  if month_in_range(r["period_key"], rng)]
    manual = fetch_manual(db_url, projects)

    wb = fill_workbook(build_workbook(), projects, month_rows, manual)
    buf = io.BytesIO()
    wb.save(buf)
    name = f"ai-sdlc-maturity_{project}_{months or 'all'}.xlsx"
    return Response(buf.getvalue(), media_type=XLSX,
                    headers={"Content-Disposition": f'attachment; filename="{name}"'})


def _check_token(token: str | None) -> None:
    expected = os.environ.get("IMPORT_TOKEN")
    if expected and token != expected:
        raise HTTPException(401, "invalid or missing import token")


_IMPORT_FORM = """
<!doctype html><meta charset="utf-8"><title>Import maturity workbook</title>
<h2>Import filled maturity workbook</h2>
<p>Download the workbook from <code>/export.xlsx</code>, fill the yellow manual
cells (Monthly: team size, costs, coverage; Quarterly: governance checklist),
then upload it here. Only manual cells are imported; auto-collected numbers are
ignored.</p>
<form action="/import/preview" method="post" enctype="multipart/form-data">
  <input type="file" name="file" accept=".xlsx" required><br><br>
  <label>Import token (if required): <input type="text" name="token"></label><br><br>
  <button type="submit">Preview changes</button>
</form>
"""


@app.get("/import", response_class=HTMLResponse)
def import_form() -> str:
    return _IMPORT_FORM


_REQUIRED_SHEETS = ("2. Projects", "3. Monthly", "4. Quarterly")


def _load_workbook(raw: bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(400, "uploaded file is not a valid .xlsx workbook")
    missing = [s for s in _REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise HTTPException(400, f"not a maturity workbook — missing sheets: {', '.join(missing)}")
    return wb


@app.post("/import/preview", response_class=HTMLResponse)
def import_preview(file: UploadFile, token: str | None = Form(default=None)) -> str:
    _check_token(token)
    wb = _load_workbook(file.file.read())
    parsed = parse_manual_inputs(wb)
    if not parsed:
        return "<p>No manual values found in the workbook. Nothing to import.</p>"
    db_url = os.environ["REPORTING_DB_URL"]
    projects = sorted({c["project"] for c in parsed})
    current = fetch_manual(db_url, projects)
    diff = diff_changes(parsed, current)
    warns = usage_warnings(parsed, fetch_auto_ai_users(db_url, projects))

    rows = "".join(
        f"<tr class={d['status']}><td>{html.escape(d['project'])}</td>"
        f"<td>{html.escape(d['period_key'])}</td><td>{html.escape(d['field'])}</td>"
        f"<td>{html.escape(str(d['old']))}</td><td>{html.escape(str(d['new']))}</td>"
        f"<td>{d['status']}</td></tr>"
        for d in diff)
    warn_html = ("<ul style='color:#b00'>" +
                 "".join(f"<li>{html.escape(w)}</li>" for w in warns) + "</ul>"
                 if warns else "")
    payload = html.escape(json.dumps(parsed))
    return f"""<!doctype html><meta charset="utf-8"><title>Preview import</title>
<h2>Preview — {len(diff)} manual value(s)</h2>{warn_html}
<table border=1 cellpadding=4><tr><th>Project</th><th>Period</th><th>Field</th>
<th>Old</th><th>New</th><th>Status</th></tr>{rows}</table>
<form action="/import/commit" method="post">
  <input type="hidden" name="changes" value="{payload}">
  <input type="hidden" name="token" value="{html.escape(token or '')}">
  <button type="submit">Confirm &amp; import</button>
</form>"""


@app.post("/import/commit", response_class=HTMLResponse)
def import_commit(changes: str = Form(...), token: str | None = Form(default=None)) -> str:
    _check_token(token)
    try:
        rows = json.loads(changes)
    except json.JSONDecodeError:
        raise HTTPException(400, "malformed changes payload")
    db_url = os.environ["REPORTING_DB_URL"]
    for c in rows:
        upsert_manual_input(db_url, c["project"], c["period_key"],
                            c["field"], c["value"], "excel-import")
    return (f"<!doctype html><meta charset='utf-8'><title>Imported</title>"
            f"<h2>Imported {len(rows)} manual value(s).</h2>"
            f"<p><a href='/import'>Import another</a></p>")
