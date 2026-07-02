"""Render the English maturity template from exporter.sheets specs.

  python -m exporter.template   # writes docs/AI SDLC Maturity.xlsx
"""
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from exporter.charts import add_charts
from exporter.sheets import SHEETS, STYLES

DST = Path("docs/AI SDLC Maturity.xlsx")


def _styler(spec: dict):
    """Precompute openpyxl style objects for one STYLES entry."""
    font = Font(name=spec.get("font", "Calibri"), sz=spec.get("size", 11),
                b=spec.get("bold", False), i=spec.get("italic", False),
                color=spec.get("color"))
    fill = (PatternFill("solid", fgColor=spec["fill"])
            if "fill" in spec else None)
    align = (Alignment(horizontal=spec.get("halign"), vertical=spec.get("valign"),
                       wrapText=spec.get("wrap", False))
             if {"halign", "valign", "wrap"} & spec.keys() else None)
    border = (Border(**{edge: Side(style=s) for edge, s in spec["border"].items()})
              if "border" in spec else None)
    numfmt = spec.get("numfmt")

    def apply(cell):
        cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        if border:
            cell.border = border
        if numfmt:
            cell.number_format = numfmt
    return apply


def build_workbook() -> openpyxl.Workbook:
    stylers = {key: _styler(spec) for key, spec in STYLES.items()}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in SHEETS:
        ws = wb.create_sheet(sheet["name"])
        if sheet.get("freeze"):
            ws.freeze_panes = sheet["freeze"]
        for col, width in sheet.get("widths", {}).items():
            ws.column_dimensions[col].width = width
        for row, height in sheet.get("heights", {}).items():
            ws.row_dimensions[row].height = height
        r0, r1 = sheet.get("data_rows", (0, -1))
        for letter, template, style in sheet.get("columns", []):
            for r in range(r0, r1 + 1):
                cell = ws[f"{letter}{r}"]
                if template:
                    cell.value = template.format(r=r)
                if style:
                    stylers[style](cell)
        for coord, value, style in sheet["cells"]:
            if isinstance(value, tuple) and value[0] == "DATE":
                value = datetime(*value[1:])
            ws[coord] = value
            if style:
                stylers[style](ws[coord])
        for merge in sheet.get("merges", []):
            ws.merge_cells(merge)
        for dv_type, formula1, sqref in sheet.get("validations", []):
            dv = DataValidation(type=dv_type, formula1=formula1, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(sqref)
        for sqref, cfvo, colors in sheet.get("color_scales", []):
            (t0, v0), (t1, v1) = cfvo[0], cfvo[-1]
            mid = {}
            if len(cfvo) == 3:
                mid = {"mid_type": cfvo[1][0], "mid_value": cfvo[1][1],
                       "mid_color": colors[1]}
            ws.conditional_formatting.add(sqref, ColorScaleRule(
                start_type=t0, start_value=v0, start_color=colors[0],
                end_type=t1, end_value=v1, end_color=colors[-1], **mid))
    add_charts(wb)
    return wb


def build() -> None:
    build_workbook().save(DST)
    print(f"wrote {DST}")


if __name__ == "__main__":
    build()
